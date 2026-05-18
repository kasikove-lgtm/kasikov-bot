"""
Бот Евгения Касикова v11.0
"""
import asyncio, os, shelve, calendar, re, logging, sys, io
from datetime import datetime, date, timedelta
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
 
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except:
    EXCEL_OK = False
 
logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout),
              logging.FileHandler("bot.log", encoding="utf-8")])
log = logging.getLogger(__name__)
 
TOKEN   = os.environ.get("BOT_TOKEN", "")
CHANNEL = os.environ.get("CHANNEL", "@kasikov_psy")
LEAD    = "https://t.me/kasikov_psy/230"
ADMINS  = ["Iozteam", "kasikovevgenii"]
 
PLATFORMS = {
    "Zoom":            "https://us04web.zoom.us/j/5806296223?pwd=Kk1taS7afUkbbdxQXnXk2FCc7Sglz4.1",
    "ВКонтакте":       "https://vk.ru/call/join/WcLhNuB_1k2NNkqmVirjcn932fkIUIgkQMQomB0kDtA",
    "Яндекс Телемост": "https://telemost.yandex.ru/j/41045386326619",
    "Google Meet":     "https://meet.google.com/xea-ubvn-sdg",
    "Teams":           "https://teams.live.com/meet/93982412886719?p=qS9poHinQUXXWNrRIp",
    "MAX":             "https://max.ru/joincall/ahmoeViSGUdzx948lSbeXgSXluuzdJW0h3HVmOepwtc",
}
DURATIONS = ["30 мин","1 час","1.5 часа","2 часа","2.5 часа","3 часа"]
BANKS     = ["Сбер","ВТБ","Альфа","Газпром","Озон"]
SOURCES   = [("📸 Instagram","ig"),("🎵 TikTok","tt"),
             ("▶️ YouTube","yt"),("💙 ВКонтакте","vk"),("🌐 Другой","other")]
PRICES = {
    "paid_first_30":3000,"30 мин":2500,"1 час":5000,
    "1.5 часа":7500,"2 часа":10000,"2.5 часа":12500,"3 часа":15000,
}
SRC_MAP = {"ig":"📸 Instagram","tt":"🎵 TikTok","yt":"▶️ YouTube",
           "vk":"💙 ВКонтакте","ref":"👥 От знакомых","other":"🌐 Другое"}
 
PAYMENT_CARD_RU = (
    "💳 *ОПЛАТА КАРТОЙ РФ / СБП*\n\n"
    "По номеру телефона СБП:\n`+7 965 763-48-79`\n"
    "Евгений Александрович К.\n\n"
    "Банки: Сбер, ВТБ, Альфа, Газпром, Озон\n\n"
    "После оплаты нажми кнопку ниже и выбери банк 👇"
)
PAYMENT_CARD_INTL = (
    "💳 *ОПЛАТА ЗАРУБЕЖНОЙ КАРТОЙ*\n\n"
    "Нажми «ПОЛУЧИТЬ РЕКВИЗИТЫ» и Евгений пришлёт реквизиты в личные сообщения.\n\n"
    "После оплаты нажми кнопку ниже 👇"
)
 
# Распорка для растяжки кнопок на полный экран
W = "\u200b" * 35
 
bot = Bot(token=TOKEN)
dp  = Dispatcher()
DB  = "kasikov_bot"
 
def db_get(k, d=None):
    with shelve.open(DB) as s: return s.get(k, d)
def db_set(k, v):
    with shelve.open(DB) as s: s[k] = v
def db_init():
    for k,v in [
        ("slots",{}),("appts",{}),("admin_chats",[]),("blocked",[]),
        ("logs",[]),("blocked_dates",[]),("users",[]),("violations",{}),
        ("regulars",[]),("states",{}),("drip",[]),("invited",[]),
        ("reviews",[]),("pending_feedback",{}),("closed_slots",{}),
        ("users_src",{}),("free_used",{}),("all_users_data",{}),
        ("feedback_timer",{}),
    ]:
        if db_get(k) is None: db_set(k, v)
db_init()
 
def get_st(uid): return db_get("states",{}).get(str(uid),{})
def set_st(uid, st):
    s = db_get("states",{}); s[str(uid)] = st; db_set("states", s)
def clr_st(uid):
    s = db_get("states",{}); s.pop(str(uid), None); db_set("states", s)
 
_fl = {}; _fc = {}
BAD = ["блять","бля","блядь","хуй","пиздец","пизда","ёбаный","ебать","ебал",
       "сука","суки","мудак","шлюха","пидор","гандон","ублюдок","долбоёб",
       "манда","жопа","нахуй","курва","дерьмо","fuck","shit","bitch","asshole","cunt"]
def has_bad(t): return any(w in t.lower() for w in BAD)
def viol(uid, tp):
    v = db_get("violations",{}); v.setdefault(str(uid),{})[tp] = v.get(str(uid),{}).get(tp,0)+1
    db_set("violations",v); return v[str(uid)][tp]
def is_flood(uid):
    now = datetime.now().timestamp()
    if now - _fl.get(uid,0) < 2:
        c = _fc.get(uid,0)+1; _fc[uid] = c
        if c > 15 and viol(uid,"spam") >= 5: asyncio.create_task(_auto_block(uid,"спам"))
        return True
    _fl[uid] = now; _fc[uid] = 0; return False
 
def is_blocked(uid): return uid in db_get("blocked",[])
def is_admin(u): return bool(u) and u.lower() in [a.lower() for a in ADMINS]
 
async def _auto_block(uid, reason):
    bl = db_get("blocked",[])
    if uid not in bl: bl.append(uid); db_set("blocked",bl)
    await notify_adm(f"🚫 *Автоблок*\nID: {uid}\nПричина: {reason}")
 
def reg_user(uid, username=None, first_name=None):
    u = db_get("users",[])
    if uid not in u: u.append(uid); db_set("users",u)
    if username or first_name:
        ud = db_get("all_users_data",{})
        ud[str(uid)] = {"username":username or "","name":first_name or "","uid":uid}
        db_set("all_users_data",ud)
 
def log_act(uid, uname, act):
    ls = db_get("logs",[]); ls.append({"uid":uid,"u":uname or "?","a":act,
        "t":datetime.now().strftime("%d.%m.%Y %H:%M:%S")})
    if len(ls) > 1000: ls = ls[-1000:]
    db_set("logs",ls)
 
async def is_sub(uid):
    try:
        m = await bot.get_chat_member(CHANNEL, uid)
        return m.status in ("member","administrator","creator")
    except Exception as e:
        log.warning(f"is_sub error {uid}: {e}")
        return False
 
async def notify_adm(text, kb=None):
    for cid in db_get("admin_chats",[]):
        try: await bot.send_message(cid, text, parse_mode="Markdown", reply_markup=kb)
        except: pass
 
def get_free_used(uid): return db_get("free_used",{}).get(str(uid), False)
def set_free_used(uid, val=True):
    fu = db_get("free_used",{}); fu[str(uid)] = val; db_set("free_used",fu)
 
def day_slots_all():
    r = []; c = datetime.strptime("10:00","%H:%M"); e = datetime.strptime("21:00","%H:%M")
    while c <= e: r.append(c.strftime("%H:%M")); c += timedelta(minutes=30)
    return r
ALL_SLOTS = day_slots_all()
MN_RU    = ["Январь","Февраль","Март","Апрель","Май","Июнь",
            "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
MN_SHORT = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]
 
def fmt_date(ds):
    try:
        d = datetime.strptime(ds,"%Y-%m-%d")
        return f"{d.day} {MN_RU[d.month-1].lower()} {d.year}"
    except: return ds
 
def get_user_src(uid):
    return db_get("users_src",{}).get(str(uid), "")
 
def _block_adj(ds, tm, dur, unblock=False):
    dm = {"30 мин":1,"1 час":2,"1.5 часа":3,"2 часа":4,"2.5 часа":5,"3 часа":6}
    n  = dm.get(dur, 2)
    if n <= 1: return
    try: sdt = datetime.strptime(f"{ds} {tm}","%Y-%m-%d %H:%M")
    except: return
    cs = db_get("closed_slots",{})
    if ds not in cs: cs[ds] = []
    for i in range(1, n):
        slot_tm = (sdt + timedelta(minutes=30*i)).strftime("%H:%M")
        if unblock:
            if slot_tm in cs[ds]: cs[ds].remove(slot_tm)
        else:
            if slot_tm not in cs[ds]: cs[ds].append(slot_tm)
    db_set("closed_slots",cs)
 
def free_slots(ds):
    slots  = db_get("slots",{}); cs = db_get("closed_slots",{})
    appts  = db_get("appts",{}); bd = db_get("blocked_dates",[])
    if ds in bd: return []
    taken  = [r["time"] for r in appts.get(ds,[])]
    closed = cs.get(ds,[])
    today  = date.today(); now = datetime.now()
    result = []
    for t in slots.get(ds,[]):
        if t in taken or t in closed: continue
        slot_dt = datetime.strptime(f"{ds} {t}","%Y-%m-%d %H:%M")
        # Клиенту показываем слоты только через 2 часа от текущего
        if datetime.strptime(ds,"%Y-%m-%d").date() == today:
            if slot_dt <= now + timedelta(hours=2): continue
        result.append(t)
    return result
 
def has_booking(uid, ds):
    return any(r["user_id"] == uid for r in db_get("appts",{}).get(ds,[]))
 
# edit_or_answer — ТОЛЬКО для админки
async def eoa(cb, text, kb=None, pm="Markdown"):
    try: await cb.message.edit_text(text, parse_mode=pm, reply_markup=kb)
    except: await cb.message.answer(text, parse_mode=pm, reply_markup=kb)
 
# ═══════════════════════════════════════════════════
# НАВИГАЦИЯ — единая логика для всех кнопок
# Глубина 1: только ↩️ НАЗАД
# Глубина 2+: ↩️ НАЗАД + ГЛАВНОЕ МЕНЮ (клиент) или МЕНЮ АДМИНА (админ)
# ═══════════════════════════════════════════════════
 
def nav_client(back_cb, depth=1):
    row = [InlineKeyboardButton(text="↩️ НАЗАД", callback_data=back_cb)]
    if depth >= 2:
        row.append(InlineKeyboardButton(text="🏠 ГЛАВНОЕ МЕНЮ", callback_data="main"))
    return [row]
 
def nav_admin(back_cb, depth=1):
    row = [InlineKeyboardButton(text="↩️ НАЗАД", callback_data=back_cb)]
    if depth >= 2:
        row.append(InlineKeyboardButton(text="🔐 МЕНЮ АДМИНА", callback_data="adm_back"))
    return [row]
 
# ═══════════════════════════════════════════════════
# ТЕКСТЫ
# ═══════════════════════════════════════════════════
 
T_ABOUT = """💼 *ОБО МНЕ*
 
Меня зовут Евгений Касиков.
 
Я не классический психолог в пиджаке с дипломом на стене. Я человек, который сам прошёл через то, с чем ты сейчас пришёл.
 
Двое детей и развод после 15 лет брака. Расставание после пяти лет отношений. Полный финансовый крах. И каждый раз казалось, что мир просто взял и перевернулся. То ощущение утром, когда просыпаешься, смотришь в потолок и думаешь: кто я теперь? Что вообще осталось?
 
Я знаю это изнутри. Не по книжкам.
 
Я из тех, кто привык добиваться результата. Кандидат в мастера спорта по плаванию, семь лет музыкальной школы. 15 лет в продажах - от мерчендайзера до директора по региону. Больше 500 собеседований, тренинги, команды. Шесть бизнесов. Флиппинг недвижимости - 175+ объектов, больше 10 лет на рынке.
 
Это дало мне навык который важен в нашей работе - быстро и точно понимать человека. Слышать не только слова, но и то, что за ними.
 
А потом всё рухнуло разом. Бизнес, отношения, темп. Деньги, статус, ориентиры. Я упал и разбился в дребезги.
 
И я не стал делать вид, что всё нормально. Пересобрал себя. В своём темпе, честно, бережно к себе, без имитации бодрости.
 
Сегодня я работаю с людьми в период расставания и развода. За плечами более 10 лет практики, более 200 часов личной и групповой терапии, более 200 реальных историй в работе с отношениями.
 
Я говорю на твоём языке. Смотрю не сверху - а рядом. Потому что я там был. И я знаю, что из этого выходят."""
 
T_HOW = """🔬 *МОИ МЕТОДЫ*
 
Скажу честно сразу - чтобы ты понимал, подходим ли мы друг другу.
 
Я не даю советов как жить. Не говорю "сделай вот так". Если ищешь именно это - я не тот специалист.
 
Я помогаю увидеть то, что ты сам не видишь. Твои паттерны, автоматические реакции, то как ты строишь отношения. Задаю вопросы - иногда неудобные. Не осуждаю - но и не сюсюкаю.
 
Решения всегда остаются за тобой. Работа идёт в живых сессиях, не в переписке. Завершить можно в любой момент.
 
Я работаю интегративно - выбираю то, что работает для конкретного человека в конкретной ситуации.
 
*НЛП*
Первые недели после расставания - хаос в голове. НЛП работает с тем, как ты воспринимаешь произошедшее. Меняет не событие, а то, как оно живёт внутри.
 
*Транзактный анализ*
Почему снова похожая ситуация, похожий партнёр, похожий финал. Смотрим из какого состояния ты живёшь в отношениях.
 
*Психология привязанности*
Почему расставание бьёт так сильно - это не слабость. Это твой тип привязанности, сформированный очень давно.
 
*EMDR*
Измена, предательство, внезапный уход - это травма. EMDR работает с тем, что застряло и не переваривается.
 
*Работа с телом*
Боль после расставания живёт не только в голове. Телесная работа помогает добраться до того, что словами не выражается.
 
*Гештальт*
Незавершённые разговоры, невысказанное. Гештальт работает в настоящем моменте - завершает прошлое.
 
*IFS - работа с частями личности*
Одна часть хочет вернуться - другая знает что нельзя. Работа с частями помогает перестать воевать с собой.
 
*Схема-терапия*
"Я недостаточно хорош", "меня всё равно бросят". Эти убеждения сформировались рано и тянут в одни и те же ситуации снова и снова.
 
*Юнгианский подход*
Когда не можешь отпустить - часто дело не в том человеке, а в том что ты видел в нём. Возвращаем это себе.
 
*Психодинамический подход*
Почему ты выбираешь именно таких людей - работаем с бессознательными паттернами.
 
*Травма-информированный подход*
Работаю аккуратно - не ломлюсь в то, к чему ты ещё не готов.
 
*Экзистенциальный подход*
После развода многие теряют не только партнёра - они теряют себя. Кризис идентичности это нормально, и из него выходят.
 
*Мужская психология*
Мужчины горюют иначе, восстанавливаются иначе. Я это учитываю."""
 
T_CONTRACT = """📋 *КАК МЫ РАБОТАЕМ*
 
*ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО*
30 минут, без обязательств. После сессии в течение 2-3 дней пришлю письменную обратную связь и план на 10 сессий.
Сессия проводится один раз. Перенос возможен если предупредил не позднее чем за 24 часа.
Не пришёл без предупреждения - следующая встреча в платном формате.
Опоздал больше чем на 10 минут без предупреждения - сессия считается состоявшейся, следующая платная.
 
*ПЛАТНЫЕ СЕССИИ*
Встречаемся в один и тот же день и время каждую неделю.
 
*ОПЛАТА*
Разовая сессия - оплата за 2-3 часа до начала в день сессии.
Пакет из 5 сессий - оплачивается полностью до первой сессии. Действует 2 месяца.
 
*ОТМЕНА И ПЕРЕНОС*
Отменить или перенести можно если предупредил не позднее чем за 24 часа.
Не предупредил - сессия оплачивается полностью.
Переносить можно не более двух раз в месяц.
 
*ОПОЗДАНИЕ*
Больше 15 минут без предупреждения - сессия оплачивается полностью.
 
*МЕЖДУ СЕССИЯМИ*
Я не консультирую в переписке. Работа - на сессиях.
Если накрыло - напиши коротко. Отвечаю в течение суток в будние дни.
Если состояние острое - найди телефон доверия или обратись в скорую.
 
*ЗАВЕРШЕНИЕ*
Завершить можно в любой момент. Предупреди за одну сессию.
 
*КОНФИДЕНЦИАЛЬНОСТЬ*
Всё что происходит на сессиях - остаётся между нами."""
 
T_PRICE = """💳 *ПЛАТНАЯ КОНСУЛЬТАЦИЯ*
 
*1-я платная встреча*
_(если на бесплатную опоздал более чем на 10 мин или отменил менее чем за 24 часа)_
Есть возможность провести встречу
30 минут - 3 000 руб.
 
*Разовая встреча*
60 минут - 5 000 руб.
_Минимальная встреча для постоянной работы - 60 минут._
 
*Пакет «Глубокая работа»*
5 встреч (5 часов) - 20 000 руб.
_экономия 5 000 руб._
 
Работаю по видео - Zoom, ВКонтакте, Яндекс Телемост, Google Meet, Teams, MAX."""
 
T_FAQ = """❓ *FAQ - ЧАСТЫЕ ВОПРОСЫ*
 
*Это конфиденциально?*
Да. Всё что происходит на сессии - остаётся между нами.
 
*Как проходит первая встреча?*
30 минут по видео. Просто поговорим - без подготовки и без обязательств. После встречи получишь письменную обратную связь по своей ситуации, моё видение как можно работать дальше и направление куда стоит смотреть в первую очередь.
 
*Можно перенести или отменить?*
Да, если предупредил минимум за 24 часа. Напиши прямо здесь в боте - нажми на свою запись в разделе 📅 Мои записи.
 
*Ты работаешь только с мужчинами?*
Нет. Около половины моих клиентов - женщины.
 
*Сколько сессий нужно?*
У всех по-разному - зависит от ситуации и запроса. После первой бесплатной встречи дам предварительные рекомендации по количеству сессий и направлениям работы.
 
*Как технически проходит сессия?*
По видеосвязи - Zoom, ВКонтакте, Яндекс Телемост, Google Meet, Teams, MAX.
 
*Я никогда не был у психолога. С чего начать?*
Просто приходи на первую бесплатную встречу. Никакой подготовки не нужно - расскажешь что происходит своими словами. Дальше разберёмся вместе.
 
*А вдруг не поможет?*
Честный ответ - я не даю гарантий. Но как правило после первой встречи становится хотя бы немного легче, и появляется понимание в каком направлении двигаться дальше.
 
*Мне сейчас очень плохо. Насколько быстро можно попасть к тебе?*
Запишись через бота - мне сразу придёт уведомление. Постараюсь подтвердить ближайший свободный слот в течение нескольких часов.
 
*Можно работать если я не в Москве?*
Да, работаю полностью онлайн. Часовой пояс согласуем.
 
*А вдруг я хочу вернуть отношения - ты поможешь?*
Да, работаю с этим. Но сначала разберёмся что за этим желанием стоит - иногда ответ меняется когда становится яснее.
 
*Как долго длится восстановление после расставания?*
У всех по-разному. Но есть вещи которые сильно ускоряют процесс - и именно с ними мы работаем.
 
*Можно прийти если партнёр не хочет на терапию?*
Да. Работаем с тобой - этого достаточно чтобы что-то начало меняться.
 
*У меня паника, не могу есть и спать. Это вообще нормально?*
Да, это типичная реакция на потерю. Именно с таким состоянием и приходят на первую встречу - не нужно ждать пока станет легче.
 
*Я уже был у психолога - не помогло.*
Понимаю. Многое зависит от метода и от того насколько специалист понимает твою конкретную ситуацию. Первая встреча покажет подходим ли мы друг другу.
 
*Онлайн - это так же эффективно как вживую?*
Да. Практика показывает что результат не отличается. Плюс не нужно никуда ехать - можно из дома в удобное время.
 
*Сколько стоит сессия?*
Нажми кнопку 💳 ПЛАТНАЯ КОНСУЛЬТАЦИЯ в главном меню - там все форматы и цены.
 
*Почему пакет из 5 сессий выгоднее?*
Глубокая работа требует времени. Пакет даёт ритм и последовательность - именно это даёт результат, а не разовые встречи.
 
*Я не знаю готов ли я вообще работать с психологом.*
Это нормально. Первая встреча ни к чему не обязывает - просто поговорим и посмотрим.
 
*Мне стыдно что не могу справиться сам.*
Справляться в одиночку со сложной ситуацией - это не сила, это лишняя нагрузка. Осуждать здесь нечего и некому.
 
*Может просто поговорить с другом?*
Можно. Но друг слышит тебя через свои фильтры и свои тревоги. На сессии ты получаешь взгляд со стороны без оценок и без чужих эмоций.
 
*Психология не противоречит моим религиозным убеждениям?*
Я не навязываю никакой картины мира. Работаем с тем что есть у тебя."""
 
T_FREE1 = """Решиться на первый шаг бывает непросто. Это нормально.
 
Я провожу бесплатную 30-минутную встречу. За это время разбираемся с тем, что сейчас происходит - и ты уходишь с ясностью и пониманием что делать дальше.
 
После встречи получишь письменный разбор с конкретными темами и рекомендациями. Он останется с тобой - независимо от того, решишь ли продолжать работу со мной."""
 
T_FREE2 = "Работаю по видео - ВКонтакте, Zoom, Яндекс Телемост, Google Meet, Teams, MAX.\n\nНайди тихое место где тебя не будут отвлекать - это важно для разговора.\n\nВыбери удобную дату 👇\n✅ - есть свободные слоты"
 
T_FEEDBACK = """Независимо от того, будешь ли ты работать со мной дальше - хочу попросить короткую обратную связь по нашей первой сессии.
 
Если откликается, ответь на несколько вопросов:
 
1. С каким состоянием ты пришёл на сессию?
2. Что из сессии было для тебя самым полезным?
3. С каким состоянием ты вышел после неё?
4. Планируешь ли продолжать работу дальше?"""
 
DRIP = {
    1:   {"text":"Гайд уже у тебя. Открывай когда будешь готов - там есть одна практика которую можно сделать прямо сегодня.", "kb":"free"},
    7:   {"text":"После расставания больно не потому что ты слабый.\n\nНейробиолог Этан Кросс: мозг воспринимает социальную боль в тех же зонах что и физическую. Когда тебе разбивают сердце - мозг регистрирует это как удар. Буквально.\n\nТы терял не просто человека. Ты терял версию себя который существовал рядом с этим человеком.\n\nБоль после расставания - это не слабость.", "kb":"channel"},
    14:  {"text":"Три вещи которые не надо делать в первый месяц. Я сам делал все три.\n\n1. Новые отношения сразу - рана переезжает, не заживает.\n2. Полный контакт с бывшим - мозгу нужна дистанция.\n3. Заливать боль чем угодно - она накапливается и потом выходит.", "kb":"channel"},
    30:  {"text":"Прошёл месяц. Как ты?\n\nЕсли чувствуешь что стало не легче а тяжелее - это сигнал. Не «возьми себя в руки». А сигнал что иногда в одиночку дольше и тяжелее.\n\nЕсли актуально 👇", "kb":"free"},
    60:  {"text":"Почему ты скучаешь по тому кто причинил тебе боль?\n\nПсихиатр Сью Джонсон: мозг привязывается не к хорошим людям. Он привязывается к знакомым.\n\nТы скучаешь не по человеку. Ты скучаешь по версии себя который верил что всё получится.\n\nКогда понимаешь это - отпускание происходит само.", "kb":"channel"},
    90:  {"text":"Почему ты снова выбираешь похожих людей?\n\nДжеффри Янг: убеждения «я недостаточно хорош», «меня всё равно бросят» формируются рано и управляют выбором партнёров.\n\nЭто не твоя вина. Но это твоя ответственность - если хочешь что-то изменить.", "kb":"channel"},
    120: {"text":"Злость после расставания - куда её деть?\n\nЛесли Гринберг: злость после расставания чаще всего вторичная эмоция. За ней прячется страх, боль, стыд.\n\nЗлость которую подавляют - разворачивается внутрь и становится депрессией. Злость которую проживают - проходит.", "kb":"channel"},
    150: {"text":"Как перестать проверять соцсети бывшего?\n\nРоберт Сапольски: мозг выделяет дофамин не когда получает награду - а когда ожидает её. Поэтому ты снова заходишь на его страницу.\n\nПервый шаг - убрать доступ. Заблокировать. Не потому что ты слабый. А потому что ты умный.", "kb":"channel"},
    180: {"text":"Разница между горем и депрессией.\n\nГоре движется. Медленно - но движется. Депрессия - стойкое ощущение пустоты больше двух недель.\n\nЕсли через два месяца не становится хотя бы немного легче - это сигнал обратиться за помощью.", "kb":"channel"},
    210: {"text":"Тревожная привязанность - почему одни страдают после расставания сильнее других.\n\nДжон Боулби: паттерн привязанности формируется в первые годы жизни.\n\nДля человека с тревожной привязанностью расставание - это не просто потеря. Это подтверждение: «я недостаточно хорош».\n\nТип привязанности не приговор - он меняется.", "kb":"channel"},
    240: {"text":"Как говорить с детьми о разводе.\n\nДжудит Валлерстайн: дети переживают не сам развод, а то как родители ведут себя во время и после него.\n\nТри вещи:\n1. «Мы решили жить отдельно. Это не твоя вина.»\n2. Никогда не делай ребёнка союзником против другого родителя.\n3. Дай ребёнку право на все чувства.", "kb":"channel"},
    270: {"text":"Когда ты готов к новым отношениям.\n\nТы можешь думать о бывшем без острой боли. Ты знаешь что пошло не так. Ты хочешь новых отношений - а не хочешь убежать от одиночества.\n\nХарвилл Хендрикс: зрелые отношения начинаются с двух целостных людей.", "kb":"free"},
    300: {"text":"Путь у каждого свой. Если захочешь пройти его с поддержкой - я здесь.\n\nПервая встреча по-прежнему бесплатно 👇", "kb":"free"},
}
 
# ═══════════════════════════════════════════════════
# МЕНЮ И КЛАВИАТУРЫ
# ═══════════════════════════════════════════════════
 
def kb_main(is_adm=False):
    rows = [
        [InlineKeyboardButton(text="🎁 ГАЙД «4 ШАГА ПОСЛЕ РАССТАВАНИЯ»",   callback_data="guide")],
        [InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО",         callback_data="free")],
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ",            callback_data="paid_info")],
        [InlineKeyboardButton(text="💼 ОБО МНЕ",                            callback_data="about")],
        [InlineKeyboardButton(text="❓ FAQ ЧАСТЫЕ ВОПРОСЫ",                 callback_data="faq")],
        [InlineKeyboardButton(text="📅 МОИ ЗАПИСИ",                         callback_data="my_appts")],
        [InlineKeyboardButton(text="📢 МОЙ TELEGRAM-КАНАЛ",
                              url=f"https://t.me/{CHANNEL.lstrip('@')}")],
    ]
    if is_adm:
        rows.append([InlineKeyboardButton(text="🔐 МЕНЮ АДМИНА", callback_data="adm_back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def kb_admin():
    """Меню админа - сразу показывает календарь + кнопки управления"""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🟡 НЕПОДТВЕРЖДЁННЫЕ ЗАПИСИ", callback_data="adm_unconf")],
        [InlineKeyboardButton(text="📋 ВСЕ ЗАПИСИ",              callback_data="adm_list")],
        [InlineKeyboardButton(text="📈 АНАЛИТИКА",               callback_data="adm_analytics")],
        [InlineKeyboardButton(text="📤 РАССЫЛКА ВСЕМ",           callback_data="adm_broadcast"),
         InlineKeyboardButton(text="📤 ПОСТ ИЗ КАНАЛА",          callback_data="adm_post")],
        [InlineKeyboardButton(text="👥 ПОСТОЯННЫЕ КЛИЕНТЫ",      callback_data="adm_regulars")],
        [InlineKeyboardButton(text="🗑 УДАЛИТЬ КЛИЕНТА",          callback_data="adm_delete_client")],
        [InlineKeyboardButton(text="🏠 МЕНЮ КЛИЕНТА",            callback_data="main")],
    ])
 
def cal_user(year, month):
    today = date.today(); cal = calendar.monthcalendar(year, month)
    bd    = db_get("blocked_dates",[]); rows = []
    rows.append([InlineKeyboardButton(text=f"📅 {MN_RU[month-1]} {year}", callback_data="noop")])
    rows.append([InlineKeyboardButton(text=d, callback_data="noop")
                 for d in ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]])
    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(text=" ", callback_data="noop"))
            else:
                d  = date(year, month, day); ds = d.strftime("%Y-%m-%d")
                if d < today:
                    row.append(InlineKeyboardButton(text="·", callback_data="noop"))
                elif ds in bd:
                    row.append(InlineKeyboardButton(text="❌", callback_data="noop"))
                elif free_slots(ds):
                    row.append(InlineKeyboardButton(text=f"✅{day}", callback_data=f"uday_{ds}"))
                else:
                    row.append(InlineKeyboardButton(text=str(day), callback_data="no_slots"))
        rows.append(row)
    pm, py = (month-1, year) if month > 1 else (12, year-1)
    nm, ny = (month+1, year) if month < 12 else (1, year+1)
    nav = []
    if date(py, pm, 1) >= today.replace(day=1):
        nav.append(InlineKeyboardButton(text="◀️", callback_data=f"ucal_{py}_{pm}"))
    else:
        nav.append(InlineKeyboardButton(text=" ", callback_data="noop"))
    nav.append(InlineKeyboardButton(text=" ", callback_data="noop"))
    nav.append(InlineKeyboardButton(text="▶️", callback_data=f"ucal_{ny}_{nm}"))
    rows.append(nav)
    rows.append([InlineKeyboardButton(text="🏠 ГЛАВНОЕ МЕНЮ", callback_data="main")])
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def cal_admin(year, month):
    today = date.today(); cal = calendar.monthcalendar(year, month)
    appts = db_get("appts",{}); slots = db_get("slots",{}); bd = db_get("blocked_dates",[]); rows = []
    rows.append([InlineKeyboardButton(text=f"🔧 {MN_RU[month-1]} {year}", callback_data="noop")])
    rows.append([InlineKeyboardButton(text=d, callback_data="noop")
                 for d in ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]])
    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(text=" ", callback_data="noop"))
            else:
                d  = date(year, month, day); ds = d.strftime("%Y-%m-%d")
                if d < today:
                    row.append(InlineKeyboardButton(text="·", callback_data=f"aday_{ds}"))
                else:
                    da = appts.get(ds,[]); is_bd = ds in bd
                    hu = any(not r.get("confirmed") for r in da)
                    hc = any(r.get("confirmed") for r in da)
                    if hu:       row.append(InlineKeyboardButton(text=f"🟡{day}", callback_data=f"aday_{ds}"))
                    elif hc:     row.append(InlineKeyboardButton(text=f"🔵{day}", callback_data=f"aday_{ds}"))
                    elif is_bd:  row.append(InlineKeyboardButton(text=f"❌{day}", callback_data=f"aday_{ds}"))
                    elif slots.get(ds): row.append(InlineKeyboardButton(text=f"✅{day}", callback_data=f"aday_{ds}"))
                    else:        row.append(InlineKeyboardButton(text=str(day), callback_data=f"aday_{ds}"))
        rows.append(row)
    pm, py = (month-1, year) if month > 1 else (12, year-1)
    nm, ny = (month+1, year) if month < 12 else (1, year+1)
    # Строка 1: навигация + записаться
    rows.append([
        InlineKeyboardButton(text="◀️", callback_data=f"acal_{py}_{pm}"),
        InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ", callback_data="adm_book_menu"),
        InlineKeyboardButton(text="▶️", callback_data=f"acal_{ny}_{nm}"),
    ])
    # Строка 2: день открыть/закрыть
    rows.append([
        InlineKeyboardButton(text="✅ ДЕНЬ ОТКРЫТЬ",   callback_data="adm_open_day_btn"),
        InlineKeyboardButton(text="❌ ДЕНЬ ЗАКРЫТЬ",   callback_data="adm_close_day_btn"),
    ])
    # Строка 3: неделю открыть/закрыть
    rows.append([
        InlineKeyboardButton(text="✅ НЕДЕЛЮ ОТКРЫТЬ", callback_data="adm_open_week_cal"),
        InlineKeyboardButton(text="❌ НЕДЕЛЮ ЗАКРЫТЬ", callback_data="adm_block_week_cal"),
    ])
    # Строка 4: месяц открыть/закрыть
    rows.append([
        InlineKeyboardButton(text="✅ МЕСЯЦ ОТКРЫТЬ",  callback_data="adm_open_month"),
        InlineKeyboardButton(text="❌ МЕСЯЦ ЗАКРЫТЬ",  callback_data="adm_block_month"),
    ])
    # Все кнопки меню админа прямо под календарём
    rows.append([InlineKeyboardButton(text="🟡 НЕПОДТВЕРЖДЁННЫЕ ЗАПИСИ", callback_data="adm_unconf")])
    rows.append([InlineKeyboardButton(text="📋 ВСЕ ЗАПИСИ",              callback_data="adm_list")])
    rows.append([InlineKeyboardButton(text="📈 АНАЛИТИКА",               callback_data="adm_analytics")])
    rows.append([
        InlineKeyboardButton(text="📤 РАССЫЛКА ВСЕМ",  callback_data="adm_broadcast"),
        InlineKeyboardButton(text="📤 ПОСТ ИЗ КАНАЛА", callback_data="adm_post"),
    ])
    rows.append([InlineKeyboardButton(text="👥 ПОСТОЯННЫЕ КЛИЕНТЫ",  callback_data="adm_regulars")])
    rows.append([InlineKeyboardButton(text="🗑 УДАЛИТЬ КЛИЕНТА",      callback_data="adm_delete_client")])
    rows.append([InlineKeyboardButton(text="🏠 МЕНЮ КЛИЕНТА",         callback_data="main")])
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def kb_month_picker(action="close"):
    today = date.today(); rows = []
    for y in [today.year, today.year + 1]:
        rows.append([InlineKeyboardButton(text=str(y), callback_data="noop")])
        for start in range(0, 12, 3):
            row = []
            for m in range(start, start+3):
                if date(y, m+1, 1) < today.replace(day=1):
                    row.append(InlineKeyboardButton(text=" ", callback_data="noop"))
                else:
                    row.append(InlineKeyboardButton(
                        text=MN_SHORT[m],
                        callback_data=f"month_pick_{action}_{y}_{m+1:02d}"))
            rows.append(row)
    rows += nav_admin("adm_cal", depth=1)
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def slots_day_admin(ds):
    slots  = db_get("slots",{}); cs = db_get("closed_slots",{})
    appts  = db_get("appts",{}); bd = db_get("blocked_dates",[])
    today  = date.today(); is_past = datetime.strptime(ds,"%Y-%m-%d").date() < today
    is_bd  = ds in bd
    open_s = slots.get(ds,[]); closed_d = cs.get(ds,[])
    all_d  = ALL_SLOTS if is_bd else sorted(set(open_s + closed_d))
    taken  = {r["time"]:r for r in appts.get(ds,[])}
    rows = []; row = []
    for t in all_d:
        if t in taken:
            em = "🔵" if taken[t].get("confirmed") else "🟡"; cb = f"aslot_{ds}_{t}"
        elif t in closed_d or is_bd:
            em = "🔴"; cb = f"aslot_blocked_{ds}_{t}"
        else:
            em = "🟢"; cb = f"aslot_{ds}_{t}"
        row.append(InlineKeyboardButton(text=f"{em}{t}", callback_data=cb))
        if len(row) == 3: rows.append(row); row = []
    if row: rows.append(row)
    if not is_past:
        rows.append([
            InlineKeyboardButton(text="✅ ОТКРЫТЬ ДЕНЬ", callback_data=f"adm_full_open_{ds}"),
            InlineKeyboardButton(text="❌ ЗАКРЫТЬ ДЕНЬ", callback_data=f"adm_full_close_{ds}"),
        ])
        rows.append([
            InlineKeyboardButton(text="➕ ДОБАВИТЬ СЛОТЫ", callback_data=f"adm_add_slots_{ds}"),
            InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬ",     callback_data=f"adm_book_slot_{ds}"),
        ])
    else:
        rows.append([InlineKeyboardButton(text="➕ ДОБАВИТЬ СЕССИЮ",
                                          callback_data=f"adm_add_session_{ds}")])
    rows += nav_admin("adm_cal", depth=1)
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def slots_menu_user(ds):
    slots  = db_get("slots",{}); cs = db_get("closed_slots",{})
    appts  = db_get("appts",{}); taken = [r["time"] for r in appts.get(ds,[])]
    closed = cs.get(ds,[]); rows = []; row = []
    for t in slots.get(ds,[]):
        if t in closed: continue
        if t in taken:
            row.append(InlineKeyboardButton(text=f"🔴{t}", callback_data="slot_taken"))
        else:
            row.append(InlineKeyboardButton(text=f"🟢{t}", callback_data=f"uslot_{ds}_{t}"))
        if len(row) == 4: rows.append(row); row = []
    if row: rows.append(row)
    rows += nav_client("main", depth=1)
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def kb_platform(back_cb="free", depth=3):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"{W}🎥 ZOOM{W}",            callback_data="plat_Zoom")],
        [InlineKeyboardButton(text=f"{W}📱 ВКОНТАКТЕ{W}",       callback_data="plat_ВКонтакте")],
        [InlineKeyboardButton(text=f"{W}💻 ЯНДЕКС ТЕЛЕМОСТ{W}", callback_data="plat_Яндекс Телемост")],
        [InlineKeyboardButton(text=f"{W}📹 GOOGLE MEET{W}",     callback_data="plat_Google Meet")],
        [InlineKeyboardButton(text=f"{W}🖥 TEAMS{W}",           callback_data="plat_Teams")],
        [InlineKeyboardButton(text=f"{W}📲 MAX{W}",             callback_data="plat_MAX")],
    ] + nav_client(back_cb, depth=depth))
 
def kb_platform_adm(back_cb="adm_back", depth=2):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🎥 ZOOM",            callback_data="plat_Zoom")],
        [InlineKeyboardButton(text="📱 ВКОНТАКТЕ",       callback_data="plat_ВКонтакте")],
        [InlineKeyboardButton(text="💻 ЯНДЕКС ТЕЛЕМОСТ", callback_data="plat_Яндекс Телемост")],
        [InlineKeyboardButton(text="📹 GOOGLE MEET",     callback_data="plat_Google Meet")],
        [InlineKeyboardButton(text="🖥 TEAMS",           callback_data="plat_Teams")],
        [InlineKeyboardButton(text="📲 MAX",             callback_data="plat_MAX")],
    ] + nav_admin(back_cb, depth=depth))
 
def kb_duration(depth=3):
    rows = [[InlineKeyboardButton(text=f"{W}{d}{W}", callback_data=f"dur_{d}")] for d in DURATIONS]
    rows += nav_client("paid_info", depth=depth)
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def kb_banks(uid, ds, tm):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=b, callback_data=f"bank_{b}_{uid}_{ds}_{tm}")] for b in BANKS
    ])
 
def get_clients_kb(step, back_cb="adm_book_menu"):
    ud = db_get("all_users_data",{}); appts = db_get("appts",{})
    seen = set(); users_data = []
    for info in ud.values():
        u = info.get("username",""); n = info.get("name","")
        if u and u not in seen:
            seen.add(u); users_data.append({"name":n,"username":u,"uid":info.get("uid",0)})
    for ds,recs in appts.items():
        for r in recs:
            u = r.get("username","")
            if u and u != "нет" and u not in seen:
                seen.add(u); users_data.append({"name":r.get("name","—"),"username":u,"uid":r.get("user_id",0)})
    rows = []
    for u in users_data[:20]:
        lbl = f"👤 {u['name']} @{u['username']}" if u['name'] else f"👤 @{u['username']}"
        rows.append([InlineKeyboardButton(text=lbl, callback_data=f"pick_client_{step}_{u['username']}")])
    rows.append([InlineKeyboardButton(text="🔍 НАЙТИ ПО ИМЕНИ", callback_data=f"search_client_{step}")])
    rows.append([InlineKeyboardButton(text="✏️ ВВЕСТИ ВРУЧНУЮ", callback_data=f"manual_client_{step}")])
    rows += nav_admin(back_cb, depth=2)
    return InlineKeyboardMarkup(inline_keyboard=rows)
 
def make_excel(d_from, d_to):
    if not EXCEL_OK: return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Записи"
    hdr = ["Дата","Время","Тип","Статус","Имя","Username","Платформа",
           "Длительность","Описание","Источник записи","Источник трафика","Оплата"]
    hf = PatternFill("solid", fgColor="2E86AB")
    for i,h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i, value=h); c.fill = hf
        c.font = Font(color="FFFFFF", bold=True); c.alignment = Alignment(horizontal="center")
    appts = db_get("appts",{}); users_src = db_get("users_src",{}); row = 2
    for ds in sorted(appts.keys()):
        if ds < d_from or ds > d_to: continue
        for r in appts[ds]:
            uid_str = str(r.get("user_id",""))
            src_traffic = users_src.get(uid_str, "")
            ws.cell(row=row,column=1,value=fmt_date(ds))
            ws.cell(row=row,column=2,value=r.get("time",""))
            ws.cell(row=row,column=3,value="Платная" if r.get("type")=="paid" else "Бесплатная")
            ws.cell(row=row,column=4,value=r.get("status",""))
            ws.cell(row=row,column=5,value=r.get("name",""))
            ws.cell(row=row,column=6,value=f"@{r.get('username','')}")
            ws.cell(row=row,column=7,value=r.get("platform",""))
            ws.cell(row=row,column=8,value=r.get("duration",""))
            ws.cell(row=row,column=9,value=r.get("desc",""))
            ws.cell(row=row,column=10,value=r.get("source","бот"))
            ws.cell(row=row,column=11,value=src_traffic)
            ws.cell(row=row,column=12,value=r.get("bank","") if r.get("paid") else "не оплачено")
            row += 1
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml+4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
 
def drip_add(uid):
    q = db_get("drip",[]); now = datetime.now()
    for day in DRIP:
        q.append({"uid":uid,"day":day,"at":(now+timedelta(days=day)).isoformat()})
    db_set("drip",q)
 
# ═══════════════════════════════════════════════════
# ФОНОВЫЕ ЗАДАЧИ
# ═══════════════════════════════════════════════════
 
async def bg_loop():
    while True:
        await asyncio.sleep(60)
        try:
            now   = datetime.now()
            appts = db_get("appts",{})
            for ds, recs in appts.items():
                changed = False
                for rec in recs:
                    try: adt = datetime.strptime(f"{ds} {rec['time']}","%Y-%m-%d %H:%M")
                    except: continue
                    diff  = (adt - now).total_seconds() / 60
                    plat  = rec.get("platform",""); link = PLATFORMS.get(plat,"")
                    tl    = "💼 Платная" if rec.get("type")=="paid" else "🆓 Бесплатная"
                    is_free = rec.get("type") != "paid"
 
                    if 59 <= diff <= 61 and not rec.get("rem_client") and rec.get("user_id"):
                        warn = (
                            "Если нужно отменить - напиши минимум за 24 часа. "
                            "Отмена позже или неявка закрывает бесплатный формат. "
                            "Следующая встреча будет возможна за 3 000 руб. / 30 минут."
                        ) if is_free else (
                            "Если нужно отменить - напиши минимум за 24 часа. "
                            "Отмена позже или неявка - сессия считается состоявшейся и оплачивается полностью."
                        )
                        try:
                            kb = InlineKeyboardMarkup(inline_keyboard=[
                                [InlineKeyboardButton(text="✅ ПОДТВЕРЖДАЮ", callback_data=f"cli_ok_{ds}_{rec['time']}")],
                                [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"cli_cancel_{ds}_{rec['time']}")],
                                [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data=f"cli_move_{ds}_{rec['time']}")],
                            ])
                            await bot.send_message(rec["user_id"],
                                f"⏰ Напоминание!\n\nЧерез час наша встреча - {rec['time']} МСК\n"
                                f"Платформа: {plat}\n{'🔗 '+link if link else ''}\n\n"
                                f"Если всё окей 😁 просьба подтвердить.\n\n{warn}",
                                reply_markup=kb)
                            rec["rem_client"] = True; changed = True
                        except: pass
 
                    if 59 <= diff <= 61 and not rec.get("rem_admin"):
                        kb2 = InlineKeyboardMarkup(inline_keyboard=[
                            [InlineKeyboardButton(text="✅ ПОДТВЕРДИТЬ", callback_data=f"adm_ok_{rec.get('user_id',0)}_{ds}_{rec['time']}")],
                            [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"adm_cncl_{rec.get('user_id',0)}_{ds}_{rec['time']}")],
                            [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data=f"adm_mv_{rec.get('user_id',0)}_{ds}_{rec['time']}")],
                        ])
                        await notify_adm(
                            f"⏰ *Через час консультация!*\n\n"
                            f"📅 {fmt_date(ds)} в {rec['time']} МСК\n{tl}\n"
                            f"👤 {rec.get('name','—')}\n📱 {plat}\n"
                            f"{'🔗 '+link if link else ''}\n"
                            f"⏱ {rec.get('duration','—')}\n🆔 @{rec.get('username','—')}",
                            kb=kb2)
                        rec["rem_admin"] = True; changed = True
 
                    if 29 <= diff <= 31 and rec.get("rem_client") and not rec.get("cli_confirmed") and rec.get("user_id"):
                        try:
                            kb = InlineKeyboardMarkup(inline_keyboard=[
                                [InlineKeyboardButton(text="✅ ПОДТВЕРЖДАЮ", callback_data=f"cli_ok_{ds}_{rec['time']}")],
                                [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"cli_cancel_{ds}_{rec['time']}")],
                                [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data=f"cli_move_{ds}_{rec['time']}")],
                            ])
                            await bot.send_message(rec["user_id"],
                                "Ты ещё не подтвердил участие. Всё в силе? 😊", reply_markup=kb)
                        except: pass
 
                    dm2 = {"30 мин":30,"1 час":60,"1.5 часа":90,"2 часа":120,"2.5 часа":150,"3 часа":180}.get(rec.get("duration","30 мин"),30)
                    past = (now-(adt+timedelta(minutes=dm2+5))).total_seconds()/60
                    if 0 <= past <= 2 and not rec.get("session_asked"):
                        kb3 = InlineKeyboardMarkup(inline_keyboard=[
                            [InlineKeyboardButton(text="✅ ДА, БЕСПЛАТНАЯ",     callback_data=f"sess_yes_free_{ds}_{rec['time']}")],
                            [InlineKeyboardButton(text="✅ ДА, ПЛАТНАЯ",         callback_data=f"sess_yes_paid_{ds}_{rec['time']}")],
                            [InlineKeyboardButton(text="❌ НЕТ, БЕСПЛАТНАЯ",    callback_data=f"sess_no_free_{ds}_{rec['time']}")],
                            [InlineKeyboardButton(text="❌ НЕТ, ПЛАТНАЯ",        callback_data=f"sess_no_paid_{ds}_{rec['time']}")],
                            [InlineKeyboardButton(text="⏰ ОПОЗДАЛ/НЕ ПРИШЁЛ", callback_data=f"sess_noshow_{ds}_{rec['time']}")],
                        ])
                        await notify_adm(
                            f"📅 {fmt_date(ds)} в {rec['time']} МСК\n{tl} | ⏱ {rec.get('duration','—')}\n"
                            f"👤 {rec.get('name','—')}\n🆔 @{rec.get('username','—')}\n\nСессия состоялась?",
                            kb=kb3)
                        rec["session_asked"] = True; changed = True
 
                if changed:
                    a = db_get("appts",{}); a[ds] = recs; db_set("appts",a)
 
            pf = db_get("pending_feedback",{})
            for key, item in list(pf.items()):
                if item.get("sent"): continue
                if now >= datetime.fromisoformat(item["remind_at"]):
                    await notify_adm(
                        f"📅 {fmt_date(item['ds'])} в {item['time']}\n"
                        f"🆓 Бесплатная | ⏱ {item.get('dur','—')}\n"
                        f"👤 {item['name']}\n🆔 @{item['username']}\n\n"
                        "Жду обратную связь и план работы, Евгений.",
                        kb=InlineKeyboardMarkup(inline_keyboard=[[
                            InlineKeyboardButton(text="📝 ОТПРАВИТЬ ОБРАТНУЮ СВЯЗЬ",
                                                 callback_data=f"send_fb_{key}")
                        ]]))
                    item["remind_at"] = (now+timedelta(hours=24)).isoformat()
            db_set("pending_feedback",pf)
 
            ft = db_get("feedback_timer",{})
            for key, item in list(ft.items()):
                if item.get("sent"): continue
                if now >= datetime.fromisoformat(item["send_at"]):
                    try:
                        await bot.send_message(item["uid"], T_FEEDBACK)
                        item["sent"] = True
                    except: pass
            db_set("feedback_timer",ft)
 
            drip = db_get("drip",[]); left = []
            for item in drip:
                if now >= datetime.fromisoformat(item["at"]):
                    d = DRIP.get(item["day"])
                    if d:
                        try:
                            kb = None
                            if d["kb"] == "free":
                                kb = InlineKeyboardMarkup(inline_keyboard=[[
                                    InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО", callback_data="free")
                                ]])
                            elif d["kb"] == "channel":
                                kb = InlineKeyboardMarkup(inline_keyboard=[[
                                    InlineKeyboardButton(text="📢 Подписаться на канал @kasikov_psy",
                                                         url=f"https://t.me/{CHANNEL.lstrip('@')}")
                                ]])
                            await bot.send_message(item["uid"], d["text"], reply_markup=kb)
                        except: pass
                else: left.append(item)
            db_set("drip",left)
 
        except Exception as e: log.error(f"bg_loop: {e}")
 
# ═══════════════════════════════════════════════════
# СТАРТ И КЛИЕНТСКИЕ ХЭНДЛЕРЫ
# ═══════════════════════════════════════════════════
 
@dp.message(Command("start","admin","stop"))
async def cmd_handler(msg: types.Message):
    uid = msg.from_user.id
    if is_blocked(uid): return
    clr_st(uid)
    reg_user(uid, msg.from_user.username, msg.from_user.first_name)
    log_act(uid, msg.from_user.username, msg.text or "cmd")
    if is_admin(msg.from_user.username):
        chats = db_get("admin_chats",[])
        if uid not in chats: chats.append(uid); db_set("admin_chats",chats)
    cmd = (msg.text or "").split()[0]
    if cmd == "/stop":
        if is_admin(msg.from_user.username):
            await msg.answer("🛑 Бот останавливается..."); await dp.stop_polling()
        return
    if cmd == "/admin":
        if is_admin(msg.from_user.username):
            today = date.today()
            await msg.answer(
                "📅 *УПРАВЛЕНИЕ ДНЯМИ*\n\n✅ свободный  🔵 записи  🟡 неподтверждённые  ❌ закрыт",
                parse_mode="Markdown", reply_markup=cal_admin(today.year, today.month))
        return
 
    # Deep link - фиксируем источник автоматически
    args = (msg.text or "").split()
    ref  = args[1] if len(args) > 1 else ""
    if ref in SRC_MAP:
        us = db_get("users_src",{}); us[str(uid)] = SRC_MAP[ref]; db_set("users_src",us)
        log_act(uid, msg.from_user.username, f"src:{ref}")
 
    # Проверяем ожидающую запись
    appts = db_get("appts",{}); today_d = date.today(); pending = None
    for ds, recs in appts.items():
        try: d = datetime.strptime(ds,"%Y-%m-%d").date()
        except: continue
        if d < today_d: continue
        for r in recs:
            if r.get("username","").lower()==(msg.from_user.username or "").lower() and r.get("user_id",0)==0:
                r["user_id"] = uid; pending = (ds,r); break
        if pending: break
    if pending:
        db_set("appts",appts); ds, r = pending
        tl   = "💼 Платная" if r.get("type")=="paid" else "🆓 Бесплатная"
        plat = r.get("platform","—"); link = PLATFORMS.get(plat,""); dur = r.get("duration","—")
        ck = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data=f"cli_move_{ds}_{r['time']}")],
            [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"cli_cancel_{ds}_{r['time']}")],
            [InlineKeyboardButton(text="📅 МОИ ЗАПИСИ",  callback_data="my_appts")],
        ])
        await msg.answer(
            f"С возвращением! 👋\n\n✅ Твоя запись уже создана:\n\n"
            f"📅 {fmt_date(ds)} в {r['time']} МСК\n{tl} | ⏱ {dur}\n📱 {plat}\n"
            f"{'🔗 '+link if link else ''}\n\nЗа час до встречи придёт напоминание.",
            reply_markup=ck)
        return
 
    is_adm = is_admin(msg.from_user.username)
    await msg.answer(
        "Хорошо что решил зайти. Это уже первый шаг.\n\nВыбери с чего начнём 👇",
        reply_markup=kb_main(is_adm=is_adm))
 
@dp.callback_query(F.data == "noop")
async def noop(cb): await cb.answer()
@dp.callback_query(F.data == "no_slots")
async def no_slots(cb): await cb.answer("На этот день слотов нет", show_alert=True)
@dp.callback_query(F.data == "slot_taken")
async def slot_taken(cb): await cb.answer("Это время занято 🔴", show_alert=True)
 
@dp.callback_query(F.data == "main")
async def go_main(cb: types.CallbackQuery):
    await cb.answer()
    is_adm = is_admin(cb.from_user.username)
    await cb.message.answer("🏠 Главное меню:", reply_markup=kb_main(is_adm=is_adm))
 
@dp.callback_query(F.data == "adm_back")
async def adm_back(cb: types.CallbackQuery):
    await cb.answer()
    today = date.today()
    await eoa(cb,
        "📅 *УПРАВЛЕНИЕ ДНЯМИ*\n\n✅ свободный  🔵 записи  🟡 неподтверждённые  ❌ закрыт",
        kb=cal_admin(today.year, today.month))
 
@dp.callback_query(F.data == "adm_menu")
async def adm_menu(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    await cb.answer()
    await eoa(cb, "🔐 *МЕНЮ АДМИНИСТРАТОРА*", kb=kb_admin())
 
@dp.callback_query(F.data == "about")
async def about(cb: types.CallbackQuery):
    await cb.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔬 МОИ МЕТОДЫ",                callback_data="how_work")],
        [InlineKeyboardButton(text="📋 КАК МЫ РАБОТАЕМ",           callback_data="contract")],
        [InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО", callback_data="free")],
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ",    callback_data="paid_info")],
    ] + nav_client("main", depth=1))
    await cb.message.answer(T_ABOUT, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data == "how_work")
async def how_work(cb: types.CallbackQuery):
    await cb.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 КАК МЫ РАБОТАЕМ",           callback_data="contract")],
        [InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО", callback_data="free")],
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ",    callback_data="paid_info")],
    ] + nav_client("about", depth=2))
    await cb.message.answer(T_HOW, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data == "contract")
async def contract_cb(cb: types.CallbackQuery):
    await cb.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔬 МОИ МЕТОДЫ",                callback_data="how_work")],
        [InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО", callback_data="free")],
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ",    callback_data="paid_info")],
    ] + nav_client("about", depth=2))
    await cb.message.answer(T_CONTRACT, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data == "faq")
async def faq(cb: types.CallbackQuery):
    await cb.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО", callback_data="free")],
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ",    callback_data="paid_info")],
    ] + nav_client("main", depth=1))
    await cb.message.answer(T_FAQ, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data == "paid_info")
async def paid_info(cb: types.CallbackQuery):
    await cb.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ", callback_data="book_paid")],
    ] + nav_client("main", depth=1))
    await cb.message.answer(T_PRICE, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data == "guide")
async def guide(cb: types.CallbackQuery):
    await cb.answer()
    log_act(cb.from_user.id, cb.from_user.username, "guide")
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📢 ПОДПИСАТЬСЯ НА КАНАЛ",
                              url=f"https://t.me/{CHANNEL.lstrip('@')}")],
        [InlineKeyboardButton(text="✅ Я ПОДПИСАЛСЯ - ДАВАЙ ГАЙД", callback_data="check_sub")],
    ] + nav_client("main", depth=1))
    await cb.message.answer(
        f"🎁 *ГАЙД «4 ШАГА ПОСЛЕ РАССТАВАНИЯ»*\n\n"
        f"30 страниц практики.\n\n"
        f"Никуда не торопись - открывай когда будешь готов, читай в своём темпе. "
        f"Гайд никуда не денется.\n\nЧтобы получить - подпишись на канал 👇{W}",
        parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data == "check_sub")
async def check_sub(cb: types.CallbackQuery):
    uid = cb.from_user.id
    if await is_sub(uid):
        log_act(uid, cb.from_user.username, "got_guide")
        drip_add(uid)
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🤝 ПЕРВАЯ ВСТРЕЧА - БЕСПЛАТНО", callback_data="free")],
        ])
        await cb.message.answer(
            f"✅ Держи гайд!\n\n"
            f"Спокойно в своём темпе - открывай когда будешь готов.\n"
            f"👉 {LEAD}\n\n"
            f"Если захочется разобрать свою ситуацию лично - первая встреча бесплатно 👇",
            reply_markup=kb)
        await cb.answer()
    else:
        await cb.answer(
            "❌ Подписка не найдена.\nПодпишись на канал и нажми кнопку снова.",
            show_alert=True)
 
@dp.callback_query(F.data == "free")
async def free_consult(cb: types.CallbackQuery):
    uid = cb.from_user.id
    if get_free_used(uid):
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ", callback_data="paid_info")],
        ] + nav_client("main", depth=1))
        await cb.answer()
        await cb.message.answer(
            "Бесплатная встреча уже была использована или не состоялась по твоей инициативе.\n\n"
            "Первая платная встреча - 30 минут - 3 000 руб.",
            reply_markup=kb)
        return
    await cb.answer()
    set_st(uid, {"type":"free"})
    log_act(uid, cb.from_user.username, "free_consult")
    await cb.message.answer(T_FREE1)
    await asyncio.sleep(0.3)
    today = date.today()
    await cb.message.answer(
        T_FREE2 + W,
        reply_markup=cal_user(today.year, today.month))
 
@dp.callback_query(F.data == "book_paid")
async def book_paid(cb: types.CallbackQuery):
    await cb.answer()
    set_st(cb.from_user.id, {"type":"paid"})
    today = date.today()
    await cb.message.answer(
        f"Выбери удобную дату 👇\n✅ - есть свободные слоты{W}",
        reply_markup=cal_user(today.year, today.month))
 
@dp.callback_query(F.data.startswith("ucal_"))
async def ucal_nav(cb: types.CallbackQuery):
    _, y, m = cb.data.split("_")
    try: await cb.message.edit_reply_markup(reply_markup=cal_user(int(y),int(m)))
    except: pass
 
@dp.callback_query(F.data.startswith("uday_"))
async def uday_sel(cb: types.CallbackQuery):
    ds = cb.data[5:]; uid = cb.from_user.id; st = get_st(uid)
    if st.get("type") != "reschedule" and has_booking(uid,ds):
        await cb.answer("У тебя уже есть запись на этот день.", show_alert=True); return
    st["date"] = ds; set_st(uid,st); await cb.answer()
    await cb.message.answer(
        f"📅 *{fmt_date(ds)}*\n\nВыбери время:\n🟢 свободно  🔴 занято{W}",
        parse_mode="Markdown", reply_markup=slots_menu_user(ds))
 
@dp.callback_query(F.data.startswith("uslot_"))
async def uslot_sel(cb: types.CallbackQuery):
    parts = cb.data.split("_"); ds, tm = parts[1], parts[2]
    uid   = cb.from_user.id; st = get_st(uid); ctype = st.get("type","free")
    await cb.answer()
    if ctype in ("free","reschedule"):
        st.update({"step":"desc","date":ds,"time":tm,"duration":"30 мин"}); set_st(uid,st)
        await cb.message.answer(
            "Отлично, записал 👍\n\n"
            "Чтобы наша встреча была для тебя максимально полезной - по желанию напиши коротко "
            "что сейчас происходит. Это поможет мне подготовиться и сразу войти в суть.\n\n"
            "Если не хочешь - просто пропусти, встретимся и разберёмся на месте.")
    else:
        st.update({"step":"duration","date":ds,"time":tm}); set_st(uid,st)
        await cb.message.answer(f"Выбери длительность встречи:{W}", reply_markup=kb_duration(depth=3))
 
@dp.callback_query(F.data.startswith("dur_"))
async def dur_sel(cb: types.CallbackQuery):
    dur = cb.data[4:]; uid = cb.from_user.id; st = get_st(uid)
    st["duration"] = dur; set_st(uid,st)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"{W}1️⃣ ПЕРВАЯ ПЛАТНАЯ ВСТРЕЧА{W}",   callback_data="session_first")],
        [InlineKeyboardButton(text=f"{W}🔄 ПОВТОРНАЯ ПЛАТНАЯ ВСТРЕЧА{W}", callback_data="session_repeat")],
    ] + nav_client("book_paid", depth=3))
    await cb.message.answer(f"Это первая или повторная платная встреча?{W}", reply_markup=kb)
 
@dp.callback_query(F.data == "session_first")
async def session_first(cb: types.CallbackQuery):
    uid = cb.from_user.id; st = get_st(uid); st["step"] = "desc"; set_st(uid,st)
    await cb.answer()
    await cb.message.answer(
        "Отлично, записал 👍\n\n"
        "Чтобы наша встреча была для тебя максимально полезной - по желанию напиши коротко "
        "что сейчас происходит. Это поможет мне подготовиться и сразу войти в суть.\n\n"
        "Если не хочешь - просто пропусти, встретимся и разберёмся на месте.")
 
@dp.callback_query(F.data == "session_repeat")
async def session_repeat(cb: types.CallbackQuery):
    uid = cb.from_user.id; st = get_st(uid)
    st["step"] = "platform"; st["desc"] = "Повторная встреча"
    st["name"] = cb.from_user.first_name or "Клиент"
    set_st(uid,st); await cb.answer()
    await cb.message.answer(
        f"Окей, разберёмся на месте. Теперь выбери платформу 👇{W}",
        reply_markup=kb_platform("book_paid", depth=3))
 
@dp.callback_query(F.data == "desc_extend")
async def desc_extend(cb: types.CallbackQuery):
    uid = cb.from_user.id; st = get_st(uid); st["step"] = "desc_retry"; set_st(uid,st)
    await cb.answer()
    await cb.message.answer("Пожалуйста, расскажи подробнее:")
 
@dp.callback_query(F.data == "desc_keep")
async def desc_keep(cb: types.CallbackQuery):
    uid = cb.from_user.id; st = get_st(uid); st["step"] = "platform"; set_st(uid,st)
    await cb.answer()
    await cb.message.answer(
        f"Окей, разберёмся на месте. Теперь выбери платформу 👇{W}",
        reply_markup=kb_platform("free", depth=3))
 
@dp.callback_query(F.data.startswith("plat_"))
async def plat_sel(cb: types.CallbackQuery):
    plat = cb.data[5:]; uid = cb.from_user.id; st = get_st(uid)
    if st.get("step") == "adm_manual_plat":
        st["platform"] = plat; st["step"] = "adm_manual_type"; set_st(uid,st); await cb.answer()
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🆓 БЕСПЛАТНАЯ", callback_data="adm_mtype_free")],
            [InlineKeyboardButton(text="💼 ПЛАТНАЯ",    callback_data="adm_mtype_paid")],
        ] + nav_admin("adm_book_menu", depth=2))
        await eoa(cb, "Тип встречи?", kb=kb); return
    st["platform"] = plat; st["step"] = "confirm"; set_st(uid,st)
    ds   = st.get("date","?"); tm = st.get("time","?"); dur = st.get("duration","30 мин")
    name = st.get("name", cb.from_user.first_name or "?")
    tl   = "💼 Платная" if st.get("type") == "paid" else "🆓 Бесплатная"
    ck   = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"{W}✅ ПОДТВЕРДИТЬ{W}",  callback_data=f"confirm_{ds}_{tm}")],
        [InlineKeyboardButton(text=f"{W}✏️ ИЗМЕНИТЬ{W}",     callback_data="reschedule_full")],
    ] + nav_client("main", depth=3))
    await cb.message.answer(
        f"Проверь данные перед подтверждением:\n\n"
        f"📅 {fmt_date(ds)} в {tm} МСК\n"
        f"{tl} - 30 минут\n👤 {name}\n📱 {plat}\n\nВсё верно?{W}",
        reply_markup=ck)
 
@dp.callback_query(F.data == "reschedule_full")
async def reschedule_full(cb: types.CallbackQuery):
    uid   = cb.from_user.id; st = get_st(uid); ctype = st.get("type","free")
    set_st(uid, {"type":ctype}); today = date.today(); await cb.answer()
    await cb.message.answer(
        f"Выбери новую дату 👇\n✅ - есть свободные слоты{W}",
        reply_markup=cal_user(today.year, today.month))
 
@dp.callback_query(F.data.startswith("confirm_"))
async def confirm_book(cb: types.CallbackQuery):
    uid = cb.from_user.id; st = get_st(uid)
    if not st: await cb.answer("Ошибка, начните заново", show_alert=True); return
    ds    = st["date"]; tm = st["time"]
    name  = st.get("name", cb.from_user.first_name or "—")
    ctype = st.get("type","free"); plat = st.get("platform","—")
    dur   = st.get("duration","30 мин"); desc = st.get("desc","—")
    uname = cb.from_user.username or "нет"
    appts = db_get("appts",{})
    if ds not in appts: appts[ds] = []
    if any(r["user_id"]==uid and r["time"]==tm for r in appts[ds]):
        await cb.answer("Ты уже записан на это время", show_alert=True); return
    appts[ds].append({
        "user_id":uid,"name":name,"time":tm,"username":uname,
        "type":ctype,"platform":plat,"duration":dur,"desc":desc,
        "source":"бот","status":"","confirmed":False,
        "rem_client":False,"rem_admin":False,"session_asked":False,
        "cli_confirmed":False,"paid":False,"bank":"",
    })
    db_set("appts",appts); _block_adj(ds,tm,dur); clr_st(uid)
    tl  = "💼 Платная" if ctype=="paid" else "🆓 Бесплатная"
    ck  = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data="reschedule")],
        [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"cancel_{ds}_{tm}")],
        [InlineKeyboardButton(text="📅 МОИ ЗАПИСИ",  callback_data="my_appts")],
    ])
    await cb.message.answer(
        f"✅ Записал тебя!\n\n"
        f"📅 {fmt_date(ds)} в {tm} МСК\n"
        f"{tl} - 30 минут\n📱 {plat}\n\n"
        f"Скоро подтвержу и пришлю ссылку для входа.",
        reply_markup=ck)
    adm_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ ПОДТВЕРДИТЬ", callback_data=f"adm_ok_{uid}_{ds}_{tm}")],
        [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"adm_cncl_{uid}_{ds}_{tm}")],
        [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data=f"adm_mv_{uid}_{ds}_{tm}")],
    ])
    await notify_adm(
        f"🔔 *Новая запись!*\n📅 {fmt_date(ds)} в {tm} МСК\n{tl} | ⏱ {dur}\n"
        f"👤 {name}\n📱 {plat}\n💬 {desc[:100]}\n🆔 @{uname} | ID: {uid}",
        kb=adm_kb)
    if ctype == "paid":
        price  = PRICES.get(dur,5000)
        pay_kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💳 ОПЛАТА КАРТОЙ РФ / СБП",   callback_data=f"pay_ru_{uid}_{ds}_{tm}")],
            [InlineKeyboardButton(text="💳 ОПЛАТА ЗАРУБЕЖНОЙ КАРТОЙ", callback_data=f"pay_intl_{uid}_{ds}_{tm}")],
        ])
        await bot.send_message(uid,
            f"💰 *Стоимость встречи: {price:,} руб.*\n\nВыбери способ оплаты:",
            parse_mode="Markdown", reply_markup=pay_kb)
 
@dp.callback_query(F.data == "my_appts")
async def my_appts(cb: types.CallbackQuery):
    uid = cb.from_user.id; appts = db_get("appts",{}); today = date.today(); rows = []
    for ds in sorted(appts.keys()):
        try: d = datetime.strptime(ds,"%Y-%m-%d").date()
        except: continue
        if d < today: continue
        for r in appts[ds]:
            if r["user_id"] != uid: continue
            tl = "💼" if r.get("type")=="paid" else "🆓"
            rows.append([InlineKeyboardButton(
                text=f"📅 {fmt_date(ds)} {r['time']} {tl}",
                callback_data=f"my_rec_{ds}_{r['time']}")])
    rows += nav_client("main", depth=1)
    if len(rows) == 1:
        await cb.message.answer("У тебя нет предстоящих записей.",
                                 reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    else:
        await cb.message.answer("📅 *Твои записи:*", parse_mode="Markdown",
                                 reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
 
@dp.callback_query(F.data.startswith("my_rec_"))
async def my_rec(cb: types.CallbackQuery):
    _, _, ds, tm = cb.data.split("_",3)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ", callback_data="reschedule")],
        [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",  callback_data=f"cancel_{ds}_{tm}")],
    ] + nav_client("my_appts", depth=2))
    await cb.message.answer(f"📅 {fmt_date(ds)} в {tm} МСК", reply_markup=kb)
 
@dp.callback_query(F.data == "reschedule")
async def reschedule(cb: types.CallbackQuery):
    set_st(cb.from_user.id,{"type":"reschedule"}); today = date.today(); await cb.answer()
    await cb.message.answer(
        f"Выбери новую дату 👇\n✅ - есть свободные слоты{W}",
        reply_markup=cal_user(today.year, today.month))
 
@dp.callback_query(F.data.startswith("cancel_"))
async def cancel_appt(cb: types.CallbackQuery):
    parts = cb.data.split("_"); ds = parts[1]; tm = parts[2]; uid = cb.from_user.id
    dur = "30 мин"; appts = db_get("appts",{})
    if ds in appts:
        for r in appts[ds]:
            if r["user_id"]==uid and r["time"]==tm: dur=r.get("duration","30 мин"); break
        appts[ds] = [r for r in appts[ds] if not (r["user_id"]==uid and r["time"]==tm)]
        db_set("appts",appts)
    _block_adj(ds,tm,dur,unblock=True)
    is_adm = is_admin(cb.from_user.username)
    await cb.message.answer("✅ Запись отменена. Если захочешь записаться снова - я здесь.",
                             reply_markup=kb_main(is_adm=is_adm))
    await notify_adm(f"❌ Клиент отменил запись\n📅 {fmt_date(ds)} в {tm}\n🆔 @{cb.from_user.username or 'нет'}")
 
@dp.callback_query(F.data.startswith("cli_ok_"))
async def cli_ok(cb: types.CallbackQuery):
    parts = cb.data.split("_"); ds = parts[2]; tm = parts[3]
    appts = db_get("appts",{})
    for r in appts.get(ds,[]):
        if r["user_id"]==cb.from_user.id and r["time"]==tm: r["cli_confirmed"] = True
    db_set("appts",appts); await cb.answer("✅ Отлично! Ждём тебя!")
    await notify_adm(f"✅ {cb.from_user.first_name} подтвердил участие\n📅 {fmt_date(ds)} в {tm}")
 
@dp.callback_query(F.data.startswith("cli_cancel_"))
async def cli_cancel(cb: types.CallbackQuery):
    parts = cb.data.split("_"); ds = parts[2]; tm = parts[3]; uid = cb.from_user.id
    dur = "30 мин"; appts = db_get("appts",{})
    if ds in appts:
        for r in appts[ds]:
            if r["user_id"]==uid and r["time"]==tm: dur=r.get("duration","30 мин"); break
        appts[ds] = [r for r in appts[ds] if not (r["user_id"]==uid and r["time"]==tm)]
        db_set("appts",appts)
    _block_adj(ds,tm,dur,unblock=True); await cb.answer()
    is_adm = is_admin(cb.from_user.username)
    await cb.message.answer("Запись отменена.", reply_markup=kb_main(is_adm=is_adm))
    await notify_adm(f"❌ Клиент отменил перед встречей\n📅 {fmt_date(ds)} в {tm}\n🆔 @{cb.from_user.username or 'нет'}")
 
@dp.callback_query(F.data.startswith("cli_move_"))
async def cli_move(cb: types.CallbackQuery):
    set_st(cb.from_user.id,{"type":"reschedule"}); today = date.today(); await cb.answer()
    await cb.message.answer(
        f"Выбери новую дату 👇\n✅ - есть свободные слоты{W}",
        reply_markup=cal_user(today.year, today.month))
    await notify_adm(f"🔄 Клиент хочет перенести\n🆔 @{cb.from_user.username or 'нет'}")
 
@dp.callback_query(F.data.startswith("pay_ru_"))
async def pay_ru(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid = int(parts[2]); ds = parts[3]; tm = parts[4]
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ ОПЛАТИЛ", callback_data=f"pay_ru_confirm_{uid}_{ds}_{tm}")
    ]])
    await cb.answer()
    await cb.message.answer(PAYMENT_CARD_RU, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data.startswith("pay_ru_confirm_"))
async def pay_ru_confirm(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid = int(parts[3]); ds = parts[4]; tm = parts[5]
    await cb.answer()
    await cb.message.answer(f"Выбери банк через который ты перевёл:{W}",
                             reply_markup=kb_banks(uid,ds,tm))
 
@dp.callback_query(F.data.startswith("bank_"))
async def bank_sel(cb: types.CallbackQuery):
    parts = cb.data.split("_"); bank=parts[1]; uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    appts = db_get("appts",{}); name="—"; dur="—"; price=5000
    for r in appts.get(ds,[]):
        if r["user_id"]==uid and r["time"]==tm:
            name=r.get("name","—"); dur=r.get("duration","—"); r["bank"]=bank
            price=PRICES.get(dur,5000); break
    db_set("appts",appts); await cb.answer(f"✅ Банк {bank} выбран")
    await cb.message.answer("✅ Информация об оплате отправлена Евгению. Ожидай подтверждения.")
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ ПОДТВЕРДИТЬ ОПЛАТУ",
                             callback_data=f"adm_confirm_pay_{uid}_{ds}_{tm}")
    ]])
    await notify_adm(
        f"💰 *Клиент оплатил!*\n\n👤 {name}\n🆔 @{cb.from_user.username or '—'}\n"
        f"📅 {fmt_date(ds)} в {tm} МСК\n⏱ {dur}\n💳 Банк: {bank}\n💰 Сумма: {price:,} руб.", kb=kb)
 
@dp.callback_query(F.data.startswith("adm_confirm_pay_"))
async def adm_confirm_pay(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts = cb.data.split("_"); uid=int(parts[3]); ds=parts[4]; tm=parts[5]
    appts = db_get("appts",{}); plat="—"; link=""; name="—"
    for r in appts.get(ds,[]):
        if r["user_id"]==uid and r["time"]==tm:
            r["paid"]=True; plat=r.get("platform","—"); link=PLATFORMS.get(plat,""); name=r.get("name","—")
    db_set("appts",appts)
    try:
        ck = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",  callback_data=f"cli_move_{ds}_{tm}")],
            [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",   callback_data=f"cli_cancel_{ds}_{tm}")],
            [InlineKeyboardButton(text="📅 МОИ ЗАПИСИ", callback_data="my_appts")],
        ])
        await bot.send_message(uid,
            f"✅ Оплата подтверждена!\n\n📅 {fmt_date(ds)} в {tm} МСК\n📱 {plat}\n"
            f"{'🔗 '+link if link else ''}\n\nЗа час до встречи придёт напоминание.",
            reply_markup=ck)
    except: pass
    await eoa(cb, cb.message.text + "\n\n✅ *Оплата подтверждена*")
 
@dp.callback_query(F.data.startswith("pay_intl_"))
async def pay_intl(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 ПОЛУЧИТЬ РЕКВИЗИТЫ", callback_data=f"intl_req_{uid}_{ds}_{tm}")],
        [InlineKeyboardButton(text="✅ ОПЛАТИЛ",            callback_data=f"intl_paid_{uid}_{ds}_{tm}")],
    ])
    await cb.answer()
    await cb.message.answer(PAYMENT_CARD_INTL, parse_mode="Markdown", reply_markup=kb)
 
@dp.callback_query(F.data.startswith("intl_req_"))
async def intl_req(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    await cb.answer("✅ Запрос отправлен Евгению")
    await cb.message.answer("✅ Евгений получил твой запрос и пришлёт реквизиты в ближайшее время.")
    appts = db_get("appts",{}); name="—"
    for r in appts.get(ds,[]):
        if r["user_id"]==uid and r["time"]==tm: name=r.get("name","—"); break
    await notify_adm(
        f"💳 *Запрос реквизитов зарубежной карты*\n\n👤 {name}\n"
        f"🆔 @{cb.from_user.username or '—'} | ID: {uid}\n"
        f"📅 {fmt_date(ds)} в {tm} МСК\n\nОтправь реквизиты клиенту в личку.")
 
@dp.callback_query(F.data.startswith("intl_paid_"))
async def intl_paid(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    appts = db_get("appts",{}); name="—"; dur="—"; price=5000
    for r in appts.get(ds,[]):
        if r["user_id"]==uid and r["time"]==tm:
            name=r.get("name","—"); dur=r.get("duration","—"); price=PRICES.get(dur,5000); break
    await cb.answer()
    await cb.message.answer("✅ Информация об оплате отправлена Евгению. Ожидай подтверждения.")
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ ПОДТВЕРДИТЬ ОПЛАТУ",
                             callback_data=f"adm_confirm_pay_{uid}_{ds}_{tm}")
    ]])
    await notify_adm(
        f"💰 *Клиент сообщил об оплате (зарубежная карта)*\n\n👤 {name}\n"
        f"🆔 @{cb.from_user.username or '—'}\n"
        f"📅 {fmt_date(ds)} в {tm} МСК\n⏱ {dur}\n💰 Сумма: {price:,} руб.", kb=kb)
 
@dp.callback_query(F.data.startswith("adm_ok_"))
async def adm_ok(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    appts = db_get("appts",{}); plat="—"; link=""
    for r in appts.get(ds,[]):
        if r["user_id"]==uid and r["time"]==tm:
            r["confirmed"]=True; plat=r.get("platform","—"); link=PLATFORMS.get(plat,"")
    db_set("appts",appts)
    try:
        ck = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",  callback_data=f"cli_move_{ds}_{tm}")],
            [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",   callback_data=f"cli_cancel_{ds}_{tm}")],
            [InlineKeyboardButton(text="📅 МОИ ЗАПИСИ", callback_data="my_appts")],
        ])
        await bot.send_message(uid,
            f"✅ Встреча подтверждена!\n\n📅 {fmt_date(ds)} в {tm} МСК\n📱 {plat}\n"
            f"{'🔗 '+link if link else ''}\n\n"
            "За час до встречи пришлю напоминание - там же попрошу подтвердить что всё в силе.",
            reply_markup=ck)
    except: pass
    adm_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",    callback_data=f"adm_cncl_{uid}_{ds}_{tm}")],
        [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",   callback_data=f"adm_mv_{uid}_{ds}_{tm}")],
        [InlineKeyboardButton(text="🔐 МЕНЮ АДМИНА", callback_data="adm_back")],
    ])
    await eoa(cb,
        f"✅ *Запись подтверждена!*\n\n📅 {fmt_date(ds)} в {tm} МСК\n👤 Клиент уведомлён.",
        kb=adm_kb)
 
@dp.callback_query(F.data.startswith("adm_cncl_"))
async def adm_cncl(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    dur="30 мин"; appts=db_get("appts",{})
    if ds in appts:
        for r in appts[ds]:
            if r["user_id"]==uid and r["time"]==tm: dur=r.get("duration","30 мин"); break
        appts[ds]=[r for r in appts[ds] if not (r["user_id"]==uid and r["time"]==tm)]
        db_set("appts",appts)
    _block_adj(ds,tm,dur,unblock=True)
    try:
        await bot.send_message(uid,"❌ К сожалению, это время не получится.\nДля новой записи нажми кнопку ниже:",
                               reply_markup=kb_main())
    except: pass
    await cb.answer("❌ Запись отменена")
    await eoa(cb, f"📅 *{fmt_date(ds)}* - запись отменена 🟢", kb=slots_day_admin(ds))
 
@dp.callback_query(F.data.startswith("adm_mv_"))
async def adm_mv(cb: types.CallbackQuery):
    parts = cb.data.split("_"); uid=int(parts[2]); ds=parts[3]; tm=parts[4]
    set_st(cb.from_user.id,{"step":"adm_move_new_date","move_uid":uid,"move_ds":ds,"move_tm":tm})
    today=date.today(); await cb.answer()
    await eoa(cb,"Выбери новую дату для переноса:",kb=cal_admin(today.year,today.month))
 
@dp.callback_query(F.data.startswith("sess_"))
async def sess_result(cb: types.CallbackQuery):
    parts=cb.data.split("_"); result=parts[1]; stype=parts[2]; ds=parts[3]; tm=parts[4]
    appts=db_get("appts",{})
    rec=next((r for r in appts.get(ds,[]) if r["time"]==tm),None)
    name=rec.get("name","—") if rec else "—"
    uname=rec.get("username","—") if rec else "—"
    dur=rec.get("duration","—") if rec else "—"
    cli_uid=rec.get("user_id",0) if rec else 0
    if result=="yes":
        if rec: rec["status"]="проведена"
        db_set("appts",appts)
        if stype=="free":
            key=f"{ds}_{tm}"; pf=db_get("pending_feedback",{})
            pf[key]={"uid":cli_uid,"ds":ds,"time":tm,"dur":dur,"name":name,"username":uname,
                     "sent":False,"remind_at":(datetime.now()+timedelta(minutes=5)).isoformat()}
            db_set("pending_feedback",pf)
            await eoa(cb,cb.message.text+"\n\n✅ Отмечено: проведена")
        else:
            kb=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📋 СКОПИРОВАТЬ (ПОСТОЯННЫЙ)",callback_data=f"reg_copy_{ds}_{tm}")],
                [InlineKeyboardButton(text="➕ НОВАЯ ЗАПИСЬ (ПОСТОЯННЫЙ)",callback_data=f"reg_new_{ds}_{tm}")],
            ]+nav_admin("adm_back",depth=1))
            await eoa(cb,
                f"✅ Платная встреча записана.\n\n📅 {fmt_date(ds)} в {tm}\n💼 | ⏱ {dur}\n"
                f"👤 {name}\n🆔 @{uname}",kb=kb)
    elif result=="noshow":
        if rec: rec["status"]="missed_no_show"
        db_set("appts",appts)
        if cli_uid: set_free_used(cli_uid,True)
        await eoa(cb,cb.message.text+"\n\n⏰ Отмечено: не пришёл. Следующая встреча - платная (30 мин / 3 000 руб.)")
    else:
        if rec: rec["status"]="не состоялась"
        db_set("appts",appts)
        sl="Бесплатная" if stype=="free" else "Платная"
        await eoa(cb,cb.message.text+f"\n\n❌ {sl} не состоялась.")
 
@dp.callback_query(F.data.startswith("send_fb_"))
async def send_fb(cb: types.CallbackQuery):
    key=cb.data[8:]
    set_st(cb.from_user.id,{"step":"adm_fb_text","fb_key":key}); await cb.answer()
    await eoa(cb,"Введи текст обратной связи и план работы.\nОн уйдёт клиенту вместе с предложением продолжить:")
 
@dp.callback_query(F.data.startswith("reg_copy_"))
async def reg_copy(cb: types.CallbackQuery):
    _,_,ds,tm=cb.data.split("_",3); appts=db_get("appts",{})
    rec=next((r for r in appts.get(ds,[]) if r["time"]==tm),None)
    if not rec: await cb.answer("Запись не найдена",show_alert=True); return
    regs=db_get("regulars",[])
    regs.append({"username":rec.get("username",""),"name":rec.get("name",""),"day":"","time":tm,"source":"авто"})
    db_set("regulars",regs); await cb.answer("✅ Добавлен в постоянные клиенты")
 
@dp.callback_query(F.data.startswith("reg_new_"))
async def reg_new(cb: types.CallbackQuery):
    _,_,ds,tm=cb.data.split("_",3); appts=db_get("appts",{})
    rec=next((r for r in appts.get(ds,[]) if r["time"]==tm),None)
    tg=rec.get("username","") if rec else ""; name=rec.get("name","") if rec else ""
    set_st(cb.from_user.id,{"step":"adm_book_date","manual_tg":tg,"manual_name":name})
    today=date.today(); await cb.answer()
    await eoa(cb,f"Записываю @{tg}. Выбери дату следующей встречи:",kb=cal_admin(today.year,today.month))
 
# ═══════════════════════════════════════════════════
# АДМИН - КАЛЕНДАРЬ
# ═══════════════════════════════════════════════════
 
@dp.callback_query(F.data.startswith("acal_"))
async def acal_nav(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    _,y,m=cb.data.split("_")
    try: await cb.message.edit_reply_markup(reply_markup=cal_admin(int(y),int(m)))
    except: pass
 
@dp.callback_query(F.data.startswith("aday_"))
async def aday_open(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ds=cb.data[5:]; uid=cb.from_user.id; st=get_st(uid); step=st.get("step","")
    today=date.today(); is_past=datetime.strptime(ds,"%Y-%m-%d").date()<today
    bd=db_get("blocked_dates",[])
 
    if step=="adm_open_day_pick":
        mode=st.get("mode","open")
        if mode=="close":
            if ds not in bd: bd.append(ds); db_set("blocked_dates",bd)
            slots=db_get("slots",{})
            if ds in slots: del slots[ds]; db_set("slots",slots)
            clr_st(uid); await cb.answer(f"❌ {fmt_date(ds)} закрыт")
            await eoa(cb,f"📅 *{fmt_date(ds)}* - день закрыт",kb=slots_day_admin(ds))
        else:
            slots=db_get("slots",{}); slots[ds]=ALL_SLOTS.copy(); db_set("slots",slots)
            if ds in bd: bd.remove(ds); db_set("blocked_dates",bd)
            clr_st(uid); await cb.answer(f"✅ {fmt_date(ds)} открыт")
            await eoa(cb,f"✅ {fmt_date(ds)} открыт - {len(ALL_SLOTS)} слотов",kb=slots_day_admin(ds))
        return
 
    if step=="adm_week_pick":
        start=datetime.strptime(ds,"%Y-%m-%d").date()
        slots=db_get("slots",{}); bd2=db_get("blocked_dates",[])
        for i in range(7):
            d2=start+timedelta(days=i); ds2=d2.strftime("%Y-%m-%d")
            slots[ds2]=ALL_SLOTS.copy()
            if ds2 in bd2: bd2.remove(ds2)
        db_set("slots",slots); db_set("blocked_dates",bd2); clr_st(uid)
        end_s=(start+timedelta(days=6)).strftime("%d.%m"); await cb.answer("✅ Неделя открыта")
        await eoa(cb,f"✅ Неделя с {fmt_date(ds)} по {end_s} открыта.",
                  kb=cal_admin(start.year,start.month)); return
 
    if step=="adm_block_week_pick":
        start=datetime.strptime(ds,"%Y-%m-%d").date(); appts=db_get("appts",{})
        conflicts=[(start+timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)
                   if appts.get((start+timedelta(days=i)).strftime("%Y-%m-%d"))]
        if conflicts:
            set_st(uid,{"step":"adm_block_week_confirm","block_week_start":ds})
            ct="\n".join(f"  {d}: {len(appts[d])} записей" for d in conflicts)
            kb=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ ЗАБЛОКИРОВАТЬ (ЗАПИСИ СОХРАНЯТСЯ)",callback_data="adm_bw_yes")],
            ]+nav_admin("adm_cal",depth=1))
            await cb.answer()
            await eoa(cb,f"⚠️ В этой неделе есть записи:\n{ct}\n\nВсё равно заблокировать?",kb=kb)
        else:
            bd2=db_get("blocked_dates",[])
            for i in range(7):
                ds2=(start+timedelta(days=i)).strftime("%Y-%m-%d")
                if ds2 not in bd2: bd2.append(ds2)
            db_set("blocked_dates",bd2); clr_st(uid)
            end_s=(start+timedelta(days=6)).strftime("%d.%m"); await cb.answer("❌ Неделя закрыта")
            await eoa(cb,f"❌ Неделя с {fmt_date(ds)} по {end_s} закрыта.",
                      kb=cal_admin(start.year,start.month))
        return
 
    if step=="adm_move_new_date":
        st["move_new_ds"]=ds; st["step"]="adm_move_new_time"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"Новая дата: *{fmt_date(ds)}*\nВведи новое время (`14:00`):"); return
 
    if step in ("adm_book_date","adm_reg_from_paid"):
        st["book_ds"]=ds; st["step"]="adm_book_slot_pick"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"📅 *{fmt_date(ds)}*\nВыбери слот для записи:",kb=slots_day_admin(ds)); return
 
    if step=="adm_excel_from":
        st["excel_from"]=ds; st["step"]="adm_excel_to"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"Начало периода: {fmt_date(ds)}\nТеперь выбери дату *конца* периода:",
                  kb=cal_admin(date.today().year,date.today().month)); return
 
    if step=="adm_excel_to":
        d_from=st.get("excel_from",""); d_to=ds; clr_st(uid); await cb.answer()
        if not EXCEL_OK: await cb.message.answer("❌ openpyxl не установлен."); return
        buf=make_excel(d_from,d_to)
        if buf:
            await cb.message.answer_document(
                BufferedInputFile(buf.read(),filename=f"записи_{d_from}_{d_to}.xlsx"),
                caption=f"📊 Записи с {fmt_date(d_from)} по {fmt_date(d_to)}")
        else: await cb.message.answer("Записей за этот период нет.")
        return
 
    if is_past:
        await cb.answer()
        await eoa(cb,f"📅 *{fmt_date(ds)}* (прошедший день)",kb=slots_day_admin(ds)); return
 
    await cb.answer()
    await eoa(cb,
        f"📅 *{fmt_date(ds)}*\n🟢 свободно  🔵 запись  🟡 неподтверждённая  🔴 закрыто",
        kb=slots_day_admin(ds))
 
@dp.callback_query(F.data.startswith("adm_full_open_"))
async def adm_full_open(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ds=cb.data[14:]
    slots=db_get("slots",{}); slots[ds]=ALL_SLOTS.copy(); db_set("slots",slots)
    bd=db_get("blocked_dates",[])
    if ds in bd: bd.remove(ds); db_set("blocked_dates",bd)
    cs=db_get("closed_slots",{})
    if ds in cs: cs[ds]=[]; db_set("closed_slots",cs)
    await cb.answer(f"✅ {fmt_date(ds)} полностью открыт")
    await eoa(cb,f"✅ {fmt_date(ds)} полностью открыт - {len(ALL_SLOTS)} слотов",kb=slots_day_admin(ds))
 
@dp.callback_query(F.data.startswith("adm_full_close_"))
async def adm_full_close(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ds=cb.data[15:]
    bd=db_get("blocked_dates",[])
    if ds not in bd: bd.append(ds); db_set("blocked_dates",bd)
    slots=db_get("slots",{})
    if ds in slots: del slots[ds]; db_set("slots",slots)
    await cb.answer(f"❌ {fmt_date(ds)} полностью закрыт")
    await eoa(cb,f"📅 *{fmt_date(ds)}* - день полностью закрыт",kb=slots_day_admin(ds))
 
@dp.callback_query(F.data.startswith("aslot_blocked_"))
async def aslot_blocked_open(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); ds=parts[2]; tm=parts[3]
    uid=cb.from_user.id; st=get_st(uid)
    if st.get("manual_tg") and st.get("step")=="adm_book_slot_pick":
        st["book_tm"]=tm; st["step"]="adm_manual_plat"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"✅ @{st['manual_tg']} - {fmt_date(ds)} в {tm}\n\nПлатформа?",
                  kb=kb_platform_adm("adm_book_menu",depth=2)); return
    kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🟢 ОТКРЫТЬ СЛОТ",     callback_data=f"adm_unblock_slot_{ds}_{tm}")],
        [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬ КЛИЕНТА",callback_data=f"adm_book_blocked_{ds}_{tm}")],
    ]+nav_admin(f"aday_{ds}",depth=1))
    await cb.answer()
    await eoa(cb,f"🔴 Слот {tm} на {fmt_date(ds)} - закрыт. Что сделать?",kb=kb)
 
@dp.callback_query(F.data.startswith("adm_unblock_slot_"))
async def adm_unblock_slot(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); ds=parts[3]; tm=parts[4]
    slots=db_get("slots",{}); cs=db_get("closed_slots",{}); bd=db_get("blocked_dates",[])
    if ds not in slots: slots[ds]=[]
    if tm not in slots[ds]: slots[ds].append(tm); slots[ds]=sorted(slots[ds])
    if ds in cs and tm in cs[ds]: cs[ds].remove(tm)
    if ds in bd: bd.remove(ds); db_set("blocked_dates",bd)
    db_set("slots",slots); db_set("closed_slots",cs)
    await cb.answer(f"🟢 Слот {tm} открыт")
    await eoa(cb,f"📅 *{fmt_date(ds)}*",kb=slots_day_admin(ds))
 
@dp.callback_query(F.data.startswith("aslot_"))
async def aslot_open(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); ds=parts[1]; tm=parts[2]
    uid=cb.from_user.id; st=get_st(uid); appts=db_get("appts",{})
    if st.get("step")=="adm_book_slot_pick":
        st["book_tm"]=tm; st["step"]="adm_manual_plat"; set_st(uid,st); await cb.answer()
        tg=st.get("manual_tg","")
        await eoa(cb,f"✅ @{tg} - {fmt_date(ds)} в {tm}\n\nПлатформа?",
                  kb=kb_platform_adm("adm_book_menu",depth=2)); return
    rec=next((r for r in appts.get(ds,[]) if r["time"]==tm),None)
    if rec:
        tl="💼 Платная" if rec.get("type")=="paid" else "🆓 Бесплатная"
        conf="✅ подтверждена" if rec.get("confirmed") else "🟡 не подтверждена"
        await cb.answer()
        kb=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ ПОДТВЕРДИТЬ",     callback_data=f"adm_ok_{rec['user_id']}_{ds}_{tm}")],
            [InlineKeyboardButton(text="❌ ОТМЕНИТЬ ЗАПИСЬ", callback_data=f"adm_cncl_{rec['user_id']}_{ds}_{tm}")],
            [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",       callback_data=f"adm_mv_{rec['user_id']}_{ds}_{tm}")],
        ]+nav_admin(f"aday_{ds}",depth=1))
        await eoa(cb,
            f"🔵 {fmt_date(ds)} в {tm}\n\n👤 {rec['name']} @{rec.get('username','')}\n"
            f"{tl} | ⏱ {rec.get('duration','—')}\n📱 {rec.get('platform','—')}\nСтатус: {conf}",kb=kb)
    else:
        in_s=tm in db_get("slots",{}).get(ds,[]); cs=db_get("closed_slots",{}).get(ds,[])
        kb=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(
                text="🔴 ЗАКРЫТЬ СЛОТ" if (in_s and tm not in cs) else "🟢 ОТКРЫТЬ СЛОТ",
                callback_data=f"adm_toggle_slot_{ds}_{tm}")],
            [InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬ КЛИЕНТА",callback_data=f"adm_book_to_{ds}_{tm}")],
        ]+nav_admin(f"aday_{ds}",depth=1))
        await cb.answer()
        status="🟢 Свободен" if in_s and tm not in cs else "🔴 Закрыт"
        await eoa(cb,f"{status} - {fmt_date(ds)} в {tm}",kb=kb)
 
@dp.callback_query(F.data.startswith("adm_toggle_slot_"))
async def adm_toggle_slot(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); ds=parts[3]; tm=parts[4]
    slots=db_get("slots",{}); cs=db_get("closed_slots",{})
    if ds not in slots: slots[ds]=[]
    if ds not in cs: cs[ds]=[]
    if tm in slots[ds] and tm not in cs[ds]:
        if tm not in cs[ds]: cs[ds].append(tm)
        await cb.answer(f"🔴 Слот {tm} закрыт")
    else:
        if tm not in slots[ds]: slots[ds].append(tm); slots[ds]=sorted(slots[ds])
        if tm in cs[ds]: cs[ds].remove(tm)
        await cb.answer(f"🟢 Слот {tm} открыт")
    db_set("slots",slots); db_set("closed_slots",cs)
    await eoa(cb,f"📅 *{fmt_date(ds)}*",kb=slots_day_admin(ds))
 
@dp.callback_query(F.data.startswith("adm_book_to_"))
async def adm_book_to(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); ds=parts[3]; tm=parts[4]
    uid=cb.from_user.id; st=get_st(uid)
    if st.get("manual_tg"):
        st["book_ds"]=ds; st["book_tm"]=tm; st["step"]="adm_manual_plat"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"✅ @{st['manual_tg']} - {fmt_date(ds)} в {tm}\n\nПлатформа?",
                  kb=kb_platform_adm("adm_book_menu",depth=2))
    else:
        set_st(uid,{"step":"adm_manual_tg","book_ds":ds,"book_tm":tm}); await cb.answer()
        await eoa(cb,"Введи username клиента в Telegram (без @):")
 
@dp.callback_query(F.data.startswith("adm_book_blocked_"))
async def adm_book_blocked(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); ds=parts[3]; tm=parts[4]
    uid=cb.from_user.id; st=get_st(uid)
    if st.get("manual_tg"):
        st["book_ds"]=ds; st["book_tm"]=tm; st["step"]="adm_manual_plat"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"✅ @{st['manual_tg']} - {fmt_date(ds)} в {tm}\n\nПлатформа?",
                  kb=kb_platform_adm("adm_book_menu",depth=2))
    else:
        set_st(uid,{"step":"adm_manual_tg","book_ds":ds,"book_tm":tm}); await cb.answer()
        await eoa(cb,"Введи username клиента в Telegram (без @):")
 
@dp.callback_query(F.data.startswith("adm_add_slots_"))
async def adm_add_slots(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ds=cb.data[14:]
    set_st(cb.from_user.id,{"step":"adm_add_slots_to","adm_date":ds}); await cb.answer()
    await eoa(cb,f"Слоты для *{fmt_date(ds)}* через запятую:\n`10:00, 11:00`")
 
@dp.callback_query(F.data.startswith("adm_book_slot_"))
async def adm_book_slot_btn(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ds=cb.data[14:]; uid=cb.from_user.id; st=get_st(uid)
    if st.get("manual_tg"):
        st["book_ds"]=ds; st["step"]="adm_book_slot_pick"; set_st(uid,st); await cb.answer()
        await eoa(cb,f"📅 *{fmt_date(ds)}*\nВыбери слот:",kb=slots_day_admin(ds))
    else:
        set_st(uid,{"step":"adm_manual_tg","book_ds":ds}); await cb.answer()
        await eoa(cb,"Введи username клиента в Telegram (без @):")
 
@dp.callback_query(F.data.startswith("adm_add_session_"))
async def adm_add_session(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ds=cb.data[16:]
    set_st(cb.from_user.id,{"step":"adm_session_name","session_ds":ds}); await cb.answer()
    await eoa(cb,f"Добавляю сессию за *{fmt_date(ds)}*.\nВведи имя клиента:")
 
@dp.callback_query(F.data=="adm_open_day_btn")
async def adm_open_day_btn(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_open_day_pick","mode":"open"}); await cb.answer()
    await eoa(cb,"Нажми на день для открытия:",kb=cal_admin(date.today().year,date.today().month))
 
@dp.callback_query(F.data=="adm_close_day_btn")
async def adm_close_day_btn(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_open_day_pick","mode":"close"}); await cb.answer()
    await eoa(cb,"Нажми на день для закрытия:",kb=cal_admin(date.today().year,date.today().month))
 
@dp.callback_query(F.data.in_({"adm_open_week","adm_open_week_cal"}))
async def adm_open_week_pick(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_week_pick"}); await cb.answer()
    await eoa(cb,"Нажми на любой день - с него откроется неделя:",kb=cal_admin(date.today().year,date.today().month))
 
@dp.callback_query(F.data.in_({"adm_block_week","adm_block_week_cal"}))
async def adm_block_week_pick(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_block_week_pick"}); await cb.answer()
    await eoa(cb,"Нажми на любой день - с него закроется неделя:",kb=cal_admin(date.today().year,date.today().month))
 
@dp.callback_query(F.data=="adm_open_month")
async def adm_open_month(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    await cb.answer()
    await eoa(cb,"Выбери месяц для открытия:",kb=kb_month_picker("open"))
 
@dp.callback_query(F.data=="adm_block_month")
async def adm_block_month(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    await cb.answer()
    await eoa(cb,"Выбери месяц для закрытия:",kb=kb_month_picker("close"))
 
@dp.callback_query(F.data.startswith("month_pick_"))
async def month_pick(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); action=parts[2]; y=int(parts[3]); m=int(parts[4])
    ms=f"{y}-{m:02d}"; dn=calendar.monthrange(y,m)[1]; appts=db_get("appts",{})
    conf=[f"{y}-{m:02d}-{d:02d}" for d in range(1,dn+1) if appts.get(f"{y}-{m:02d}-{d:02d}")]
    bd=db_get("blocked_dates",[]); slots=db_get("slots",{})
    if action=="open":
        for d in range(1,dn+1):
            ds=f"{y}-{m:02d}-{d:02d}"
            if ds in bd: bd.remove(ds)
            slots[ds]=ALL_SLOTS.copy()
        db_set("blocked_dates",bd); db_set("slots",slots)
        await cb.answer(f"✅ {MN_RU[m-1]} {y} открыт")
        await eoa(cb,f"✅ {MN_RU[m-1]} {y} открыт.",kb=kb_admin())
    else:
        if conf:
            set_st(cb.from_user.id,{"step":"adm_bm_confirm","block_month":ms})
            ct="\n".join(f"  {ds}: {len(appts[ds])} записей" for ds in conf)
            kb=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✅ ЗАКРЫТЬ (ЗАПИСИ СОХРАНЯТСЯ)",callback_data="adm_bm_yes")],
            ]+nav_admin("adm_cal",depth=1))
            await eoa(cb,f"⚠️ В {MN_RU[m-1]} {y} есть записи:\n{ct}\n\nВсё равно закрыть?",kb=kb)
        else:
            for d in range(1,dn+1):
                ds=f"{y}-{m:02d}-{d:02d}"
                if ds not in bd: bd.append(ds)
            db_set("blocked_dates",bd); await cb.answer(f"❌ {MN_RU[m-1]} {y} закрыт")
            await eoa(cb,f"❌ {MN_RU[m-1]} {y} закрыт.",kb=kb_admin())
 
@dp.callback_query(F.data=="adm_bm_yes")
async def adm_bm_yes(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    uid=cb.from_user.id; st=get_st(uid); ms=st.get("block_month","")
    if not ms: return
    y,m=int(ms.split("-")[0]),int(ms.split("-")[1]); dn=calendar.monthrange(y,m)[1]
    bd=db_get("blocked_dates",[])
    for d in range(1,dn+1):
        ds=f"{y}-{m:02d}-{d:02d}"
        if ds not in bd: bd.append(ds)
    db_set("blocked_dates",bd); clr_st(uid)
    await eoa(cb,f"❌ {MN_RU[m-1]} {y} закрыт (записи сохранены).",kb=kb_admin())
 
@dp.callback_query(F.data=="adm_bw_yes")
async def adm_bw_yes(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    uid=cb.from_user.id; st=get_st(uid); ds=st.get("block_week_start","")
    if not ds: return
    start=datetime.strptime(ds,"%Y-%m-%d").date(); bd=db_get("blocked_dates",[])
    for i in range(7):
        ds2=(start+timedelta(days=i)).strftime("%Y-%m-%d")
        if ds2 not in bd: bd.append(ds2)
    db_set("blocked_dates",bd); clr_st(uid)
    end_s=(start+timedelta(days=6)).strftime("%d.%m")
    await eoa(cb,f"❌ Неделя с {fmt_date(ds)} по {end_s} закрыта (записи сохранены).",
              kb=cal_admin(start.year,start.month))
 
# ═══════════════════════════════════════════════════
# АДМИН - МЕНЮ ЗАПИСИ
# ═══════════════════════════════════════════════════
 
@dp.callback_query(F.data=="adm_book_menu")
async def adm_book_menu(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ НОВЫЙ - БЕСПЛАТНАЯ", callback_data="adm_book_new_free")],
        [InlineKeyboardButton(text="✏️ НОВЫЙ - ПЛАТНАЯ",    callback_data="adm_book_new_paid")],
        [InlineKeyboardButton(text="✏️ ПОСТОЯННЫЙ",         callback_data="adm_book_reg")],
    ]+nav_admin("adm_back",depth=1))
    await cb.answer(); await eoa(cb,"Выбери тип записи:",kb=kb)
 
@dp.callback_query(F.data=="adm_book_new_free")
async def adm_book_new_free(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    await cb.answer()
    await eoa(cb,"Откуда клиент?",kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=lbl,callback_data=f"adm_src_book_{code}_free")]
        for lbl,code in SOURCES
    ]+nav_admin("adm_book_menu",depth=2)))
 
@dp.callback_query(F.data=="adm_book_new_paid")
async def adm_book_new_paid(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    await cb.answer()
    await eoa(cb,"Откуда клиент?",kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=lbl,callback_data=f"adm_src_book_{code}_paid")]
        for lbl,code in SOURCES
    ]+nav_admin("adm_book_menu",depth=2)))
 
@dp.callback_query(F.data.startswith("adm_src_book_"))
async def adm_src_book(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); code=parts[3]; forced=parts[4]
    uid=cb.from_user.id; set_st(uid,{"forced_type":forced,"source":code}); await cb.answer()
    await eoa(cb,"Выбери клиента:",kb=get_clients_kb(forced,"adm_book_new_free" if forced=="free" else "adm_book_new_paid"))
 
@dp.callback_query(F.data=="adm_book_reg")
async def adm_book_reg(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    regs=db_get("regulars",[])
    if not regs:
        kb=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✏️ ВВЕСТИ ВРУЧНУЮ",callback_data="manual_client_reg")],
        ]+nav_admin("adm_book_menu",depth=2))
        await cb.answer(); await eoa(cb,"Постоянных клиентов нет.",kb=kb); return
    rows=[[InlineKeyboardButton(
        text=f"👤 {r.get('name','')} @{r['username']}",
        callback_data=f"adm_reg_pick_{r['username']}")] for r in regs[:20]]
    rows+=[[InlineKeyboardButton(text="🔍 НАЙТИ ПО ИМЕНИ",callback_data="search_client_reg")]]
    rows+=[[InlineKeyboardButton(text="✏️ ВВЕСТИ ВРУЧНУЮ",callback_data="manual_client_reg")]]
    rows+=nav_admin("adm_book_menu",depth=2)
    await cb.answer(); await eoa(cb,"Выбери постоянного клиента:",kb=InlineKeyboardMarkup(inline_keyboard=rows))
 
@dp.callback_query(F.data.startswith("pick_client_"))
async def pick_client(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_",3); step=parts[2]; uname=parts[3]
    uid=cb.from_user.id; ud=db_get("all_users_data",{})
    client_uid=0; client_name=uname
    for info in ud.values():
        if info.get("username","").lower()==uname.lower():
            client_uid=info.get("uid",0); client_name=info.get("name",uname); break
    if not client_uid:
        for ds2,recs in db_get("appts",{}).items():
            for r in recs:
                if r.get("username","").lower()==uname.lower():
                    client_uid=r.get("user_id",0); client_name=r.get("name",uname); break
            if client_uid: break
    forced="free" if step in("free","reg") else "paid"
    st=get_st(uid)
    set_st(uid,{**st,"step":"adm_book_date","manual_tg":uname,
                "manual_name":client_name,"manual_uid":client_uid,"forced_type":forced})
    today=date.today(); await cb.answer()
    await eoa(cb,f"✅ Клиент: *{client_name}* @{uname}\nВыбери дату:",kb=cal_admin(today.year,today.month))
 
@dp.callback_query(F.data.startswith("manual_client_"))
async def manual_client(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    step=cb.data[14:]; forced="free" if step=="free" else "paid"
    set_st(cb.from_user.id,{"step":"adm_manual_tg","forced_type":forced}); await cb.answer()
    await eoa(cb,"Введи username клиента в Telegram (без @):")
 
@dp.callback_query(F.data.startswith("search_client_"))
async def search_client(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    step=cb.data[14:]; forced="free" if step=="free" else "paid"
    set_st(cb.from_user.id,{"step":"adm_search_client","forced_type":forced}); await cb.answer()
    await eoa(cb,"Введи часть имени или username для поиска:")
 
@dp.callback_query(F.data.startswith("adm_reg_pick_"))
async def adm_reg_pick(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    uname=cb.data[13:]; regs=db_get("regulars",[])
    rec=next((r for r in regs if r["username"]==uname),None)
    if not rec: await cb.answer("Не найден",show_alert=True); return
    set_st(cb.from_user.id,{"step":"adm_book_date","manual_tg":uname,
                             "manual_name":rec.get("name",""),"forced_type":"paid"})
    today=date.today(); await cb.answer()
    await eoa(cb,f"Записываю @{uname}. Выбери дату:",kb=cal_admin(today.year,today.month))
 
# ── АНАЛИТИКА ─────────────────────────────────────
 
@dp.callback_query(F.data=="adm_analytics")
async def adm_analytics(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 СТАТИСТИКА",          callback_data="adm_stats")],
        [InlineKeyboardButton(text="📊 EXCEL ОТЧЁТ",         callback_data="adm_excel")],
        [InlineKeyboardButton(text="📋 ПРОШЕДШИЕ БЕСПЛАТНЫЕ",callback_data="adm_past_free")],
        [InlineKeyboardButton(text="⭐️ ОТЗЫВЫ",              callback_data="adm_reviews")],
        [InlineKeyboardButton(text="🚫 ЗАБЛОКИРОВАННЫЕ",     callback_data="adm_blocked")],
        [InlineKeyboardButton(text="📊 ИСТОРИЯ / ЛОГИ",      callback_data="adm_logs")],
    ]+nav_admin("adm_back",depth=1))
    await cb.answer(); await eoa(cb,"📈 *АНАЛИТИКА*",kb=kb)
 
@dp.callback_query(F.data=="adm_stats")
async def adm_stats(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    appts=db_get("appts",{}); logs=db_get("logs",[]); now=datetime.now()
    total=fc=pc=done=miss=0; srcs={}
    for ds,recs in appts.items():
        try: d=datetime.strptime(ds,"%Y-%m-%d")
        except: continue
        if d.month==now.month and d.year==now.year:
            for r in recs:
                total+=1
                if r.get("type")=="paid": pc+=1
                else: fc+=1
                if r.get("status")=="проведена": done+=1
                if r.get("status")=="не состоялась": miss+=1
                s=r.get("source","бот"); srcs[s]=srcs.get(s,0)+1
    guides=sum(1 for l in logs if l.get("a")=="got_guide"); users=len(db_get("users",[]))
    us=db_get("users_src",{}); sl={}
    for v in us.values(): sl[v]=sl.get(v,0)+1
    st_text="\n".join(f"  {k}: {v}" for k,v in srcs.items()) or "  нет данных"
    sl_text="\n".join(f"  {k}: {v}" for k,v in sl.items()) or "  нет данных"
    text=(
        f"📊 *СТАТИСТИКА {now.strftime('%B %Y')}:*\n\n"
        f"📅 Всего: {total}\n🆓 Бесплатных: {fc}\n💼 Платных: {pc}\n"
        f"✅ Проведено: {done}\n❌ Не состоялось: {miss}\n"
        f"📄 Гайдов: {guides}\n👥 Пользователей: {users}\n\n"
        f"*Источники записей:*\n{st_text}\n\n"
        f"*Откуда узнали:*\n{sl_text}"
    )
    await eoa(cb,text,kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_analytics",depth=2)))
 
@dp.callback_query(F.data=="adm_excel")
async def adm_excel(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 1 ДЕНЬ",   callback_data="excel_1d")],
        [InlineKeyboardButton(text="📅 1 НЕДЕЛЯ", callback_data="excel_1w")],
        [InlineKeyboardButton(text="📅 1 МЕСЯЦ",  callback_data="excel_1m")],
        [InlineKeyboardButton(text="📅 ПЕРИОД",   callback_data="excel_period")],
    ]+nav_admin("adm_analytics",depth=2))
    await cb.answer(); await eoa(cb,"Выбери период для отчёта:",kb=kb)
 
@dp.callback_query(F.data.in_({"excel_1d","excel_1w","excel_1m"}))
async def excel_quick(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    today=date.today()
    if cb.data=="excel_1d": d_from=d_to=today.strftime("%Y-%m-%d")
    elif cb.data=="excel_1w": d_from=(today-timedelta(days=7)).strftime("%Y-%m-%d"); d_to=today.strftime("%Y-%m-%d")
    else: d_from=today.replace(day=1).strftime("%Y-%m-%d"); d_to=today.strftime("%Y-%m-%d")
    await cb.answer()
    if not EXCEL_OK: await cb.message.answer("❌ openpyxl не установлен."); return
    buf=make_excel(d_from,d_to)
    if buf:
        await cb.message.answer_document(
            BufferedInputFile(buf.read(),filename=f"записи_{d_from}_{d_to}.xlsx"),
            caption=f"📊 Записи с {fmt_date(d_from)} по {fmt_date(d_to)}")
    else: await cb.message.answer("Записей за этот период нет.")
 
@dp.callback_query(F.data=="excel_period")
async def excel_period(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_excel_from"}); await cb.answer()
    await eoa(cb,"Выбери дату *начала* периода:",kb=cal_admin(date.today().year,date.today().month))
 
@dp.callback_query(F.data=="adm_past_free")
async def adm_past_free(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    appts=db_get("appts",{}); today=date.today(); week_ago=today-timedelta(days=7); rows=[]
    for ds in sorted(appts.keys(),reverse=True):
        try: d=datetime.strptime(ds,"%Y-%m-%d").date()
        except: continue
        if d>today or d<week_ago: continue
        for r in appts[ds]:
            if r.get("type")!="free": continue
            s=r.get("status",""); ic="✅" if s=="проведена" else ("❌" if s=="не состоялась" else "⏳")
            rows.append([InlineKeyboardButton(
                text=f"{ic} {fmt_date(ds)} {r['time']} - {r['name']}",
                callback_data=f"aslot_{ds}_{r['time']}")])
    if not rows:
        await eoa(cb,"Прошедших бесплатных за неделю нет.",
                  kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_analytics",depth=2))); return
    rows+=nav_admin("adm_analytics",depth=2)
    await eoa(cb,"📋 *ПРОШЕДШИЕ БЕСПЛАТНЫЕ (7 дней):*",kb=InlineKeyboardMarkup(inline_keyboard=rows))
 
@dp.callback_query(F.data=="adm_reviews")
async def adm_reviews(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    reviews=db_get("reviews",[])
    nav=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_analytics",depth=2))
    if not reviews: await eoa(cb,"Отзывов пока нет.",kb=nav); return
    text="⭐️ *ОТЗЫВЫ:*\n\n"
    for r in reviews[-10:]: text+=f"👤 {r.get('name','—')} | {'⭐️'*r.get('stars',5)}\n{r.get('text','')}\n\n"
    await eoa(cb,text,kb=nav)
 
@dp.callback_query(F.data=="adm_blocked")
async def adm_blocked(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    bl=db_get("blocked",[])
    nav=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_analytics",depth=2))
    if not bl: await eoa(cb,"Заблокированных нет.",kb=nav); return
    kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔓 РАЗБЛОКИРОВАТЬ",callback_data="adm_unblock_ask")]
    ]+nav_admin("adm_analytics",depth=2))
    await eoa(cb,"🚫 *ЗАБЛОКИРОВАННЫЕ:*\n\n"+"\n".join(str(u) for u in bl),kb=kb)
 
@dp.callback_query(F.data=="adm_unblock_ask")
async def adm_unblock_ask(cb):
    set_st(cb.from_user.id,{"step":"adm_unblock"}); await cb.answer()
    await eoa(cb,"Введи ID для разблокировки:")
 
@dp.callback_query(F.data=="adm_logs")
async def adm_logs(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    logs=db_get("logs",[]); text="📊 *ПОСЛЕДНИЕ 20 ДЕЙСТВИЙ:*\n\n"
    for e in logs[-20:]: text+=f"🕐 {e['t']} @{e['u']}\n▶️ {e['a']}\n\n"
    await eoa(cb,text if logs else "Логов нет.",
              kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_analytics",depth=2)))
 
@dp.callback_query(F.data=="adm_unconf")
async def adm_unconf(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    appts=db_get("appts",{}); today=date.today(); rows=[]
    for ds in sorted(appts.keys()):
        try: d=datetime.strptime(ds,"%Y-%m-%d").date()
        except: continue
        if d<today: continue
        for r in appts[ds]:
            if not r.get("confirmed"):
                tl="💼" if r.get("type")=="paid" else "🆓"
                rows.append([InlineKeyboardButton(
                    text=f"🟡 {fmt_date(ds)} {r['time']} {tl} - {r['name']}",
                    callback_data=f"aslot_{ds}_{r['time']}")])
    if not rows:
        await eoa(cb,"Неподтверждённых нет.",kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_back",depth=1))); return
    rows+=nav_admin("adm_back",depth=1)
    await eoa(cb,"🟡 *НЕПОДТВЕРЖДЁННЫЕ:*",kb=InlineKeyboardMarkup(inline_keyboard=rows))
 
@dp.callback_query(F.data=="adm_list")
async def adm_list(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    appts=db_get("appts",{}); today=date.today(); text="📋 *ПРЕДСТОЯЩИЕ ЗАПИСИ:*\n\n"; found=False
    for ds in sorted(appts.keys()):
        try: d=datetime.strptime(ds,"%Y-%m-%d").date()
        except: continue
        if d<today: continue
        for r in appts[ds]:
            tl="💼" if r.get("type")=="paid" else "🆓"; co="✅" if r.get("confirmed") else "🟡"
            text+=(f"{co}{tl} *{fmt_date(ds)}* {r['time']} ⏱{r.get('duration','—')}\n"
                   f"👤 {r['name']} @{r.get('username','')}\n📱 {r.get('platform','—')}\n\n"); found=True
    await eoa(cb,text if found else "📭 Записей нет.",
              kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_back",depth=1)))
 
@dp.callback_query(F.data=="adm_regulars")
async def adm_regulars(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    regs=db_get("regulars",[]); text="👥 *ПОСТОЯННЫЕ КЛИЕНТЫ:*\n\n"
    text+="\n".join(f"@{r['username']} - {r['name']} | {r.get('day','')} {r.get('time','')}" for r in regs) if regs else "Нет."
    kb=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ ДОБАВИТЬ",callback_data="adm_reg_add")],
        [InlineKeyboardButton(text="🗑 УДАЛИТЬ",  callback_data="adm_reg_del")],
    ]+nav_admin("adm_back",depth=1))
    await eoa(cb,text,kb=kb)
 
@dp.callback_query(F.data=="adm_reg_add")
async def adm_reg_add(cb):
    set_st(cb.from_user.id,{"step":"adm_reg_username"}); await cb.answer()
    await eoa(cb,"Введи username клиента (без @):")
 
@dp.callback_query(F.data=="adm_reg_del")
async def adm_reg_del_btn(cb):
    set_st(cb.from_user.id,{"step":"adm_reg_del_name"}); await cb.answer()
    await eoa(cb,"Введи username для удаления (без @):")
 
@dp.callback_query(F.data.startswith("regday_"))
async def regday(cb: types.CallbackQuery):
    day=cb.data[7:]; uid=cb.from_user.id; st=get_st(uid)
    st["reg_day"]=day; st["step"]="adm_reg_time"; set_st(uid,st); await cb.answer()
    await eoa(cb,f"День: *{day}*\nВведи время (`15:00`):")
 
@dp.callback_query(F.data=="adm_broadcast")
async def adm_broadcast(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_broadcast_text"}); await cb.answer()
    await eoa(cb,"Введи текст рассылки:",
              kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_back",depth=1)))
 
@dp.callback_query(F.data=="adm_post")
async def adm_post(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    set_st(cb.from_user.id,{"step":"adm_post_link"}); await cb.answer()
    await eoa(cb,"Вставь ссылку на пост из канала:")
 
@dp.callback_query(F.data=="adm_delete_client")
async def adm_delete_client(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ud=db_get("all_users_data",{}); regs=db_get("regulars",[])
    seen=set(); rows=[]
    for info in ud.values():
        u=info.get("username",""); n=info.get("name","")
        if u and u not in seen:
            seen.add(u); rows.append([InlineKeyboardButton(
                text=f"🗑 {n} @{u}" if n else f"🗑 @{u}",callback_data=f"del_client_{u}")])
    for r in regs:
        u=r.get("username","")
        if u and u not in seen:
            seen.add(u); rows.append([InlineKeyboardButton(
                text=f"🗑 {r.get('name','')} @{u}",callback_data=f"del_client_{u}")])
    if not rows:
        await eoa(cb,"Клиентов в базе нет.",
                  kb=InlineKeyboardMarkup(inline_keyboard=nav_admin("adm_back",depth=1))); return
    rows+=nav_admin("adm_back",depth=1)
    await cb.answer(); await eoa(cb,"Выбери клиента для удаления:",kb=InlineKeyboardMarkup(inline_keyboard=rows))
 
@dp.callback_query(F.data.startswith("del_client_"))
async def del_client_confirm(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    if cb.data.startswith("del_client_yes_"):
        uname=cb.data[15:]
        ud=db_get("all_users_data",{})
        to_del=[k for k,v in ud.items() if v.get("username","").lower()==uname.lower()]
        for k in to_del: del ud[k]
        db_set("all_users_data",ud)
        regs=db_get("regulars",[])
        db_set("regulars",[r for r in regs if r.get("username","").lower()!=uname.lower()])
        us=db_get("users_src",{})
        to_del2=[k for k,v in us.items() if v==uname]
        for k in to_del2: del us[k]
        db_set("users_src",us)
        await cb.answer(f"✅ @{uname} удалён")
        today=date.today()
        await eoa(cb,f"✅ Клиент @{uname} удалён из всех баз.",
                  kb=cal_admin(today.year,today.month))
    else:
        uname=cb.data[11:]
        kb=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"✅ ДА, УДАЛИТЬ @{uname}",callback_data=f"del_client_yes_{uname}")],
            [InlineKeyboardButton(text="❌ ОТМЕНА",callback_data="adm_delete_client")],
        ])
        await cb.answer(); await eoa(cb,f"Удалить @{uname} из всех баз?",kb=kb)
 
@dp.callback_query(F.data.startswith("adm_mtype_"))
async def adm_mtype(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    ctype="paid" if "paid" in cb.data else "free"
    uid=cb.from_user.id; st=get_st(uid)
    ds=st.get("book_ds","?"); tm=st.get("book_tm","?"); name=st.get("manual_name","?")
    tg=st.get("manual_tg","?"); plat=st.get("platform","?"); dur=st.get("duration","—")
    forced=st.get("forced_type")
    if forced: ctype=forced
    tl="💼 Платная" if ctype=="paid" else "🆓 Бесплатная"
    client_uid=st.get("manual_uid",0)
    if not client_uid:
        ud=db_get("all_users_data",{})
        for info in ud.values():
            if info.get("username","").lower()==tg.lower(): client_uid=info.get("uid",0); break
        if not client_uid:
            for ds2,recs in db_get("appts",{}).items():
                for r in recs:
                    if r.get("username","").lower()==tg.lower() and r.get("user_id",0):
                        client_uid=r["user_id"]; break
                if client_uid: break
    appts=db_get("appts",{})
    if ds not in appts: appts[ds]=[]
    appts[ds].append({
        "user_id":client_uid,"name":name,"time":tm,"username":tg,
        "type":ctype,"platform":plat,"duration":dur,"desc":"Ручная запись",
        "source":st.get("source","вручную"),"status":"","confirmed":True,
        "rem_client":False,"rem_admin":False,"session_asked":False,
        "cli_confirmed":False,"paid":False,"bank":"",
    })
    db_set("appts",appts); clr_st(uid)
    today=date.today()
    await eoa(cb,f"✅ Записан вручную!\n\n📅 {fmt_date(ds)} в {tm}\n{tl}\n👤 {name} @{tg}\n📱 {plat}",
              kb=cal_admin(today.year,today.month))
    await notify_adm(f"✏️ *Ручная запись*\n📅 {fmt_date(ds)} в {tm}\n{tl}\n👤 {name} @{tg}")
    if client_uid:
        link=PLATFORMS.get(plat,"")
        ck=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 ПЕРЕНЕСТИ",  callback_data=f"cli_move_{ds}_{tm}")],
            [InlineKeyboardButton(text="❌ ОТМЕНИТЬ",   callback_data=f"cli_cancel_{ds}_{tm}")],
            [InlineKeyboardButton(text="📅 МОИ ЗАПИСИ", callback_data="my_appts")],
        ])
        try:
            await bot.send_message(client_uid,
                f"✅ Встреча подтверждена!\n\n📅 {fmt_date(ds)} в {tm} МСК\n{tl} | ⏱ {dur}\n📱 {plat}\n"
                f"{'🔗 '+link if link else ''}\n\nЗа час до встречи придёт напоминание.",reply_markup=ck)
        except: pass
        if ctype=="paid" and dur!="—":
            price=PRICES.get(dur,5000)
            pay_kb=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="💳 ОПЛАТА КАРТОЙ РФ / СБП",   callback_data=f"pay_ru_{client_uid}_{ds}_{tm}")],
                [InlineKeyboardButton(text="💳 ОПЛАТА ЗАРУБЕЖНОЙ КАРТОЙ", callback_data=f"pay_intl_{client_uid}_{ds}_{tm}")],
            ])
            try: await bot.send_message(client_uid,
                    f"💰 *Стоимость встречи: {price:,} руб.*\n\nВыбери способ оплаты:",
                    parse_mode="Markdown",reply_markup=pay_kb)
            except: pass
    else:
        try: me=await bot.get_me(); link2=f"https://t.me/{me.username}"
        except: link2="https://t.me/kasikov_bot"
        await notify_adm(f"⚠️ Клиент @{tg} не найден в боте.\nОтправь ему ссылку:\n{link2}")
 
@dp.callback_query(F.data.startswith("stype_"))
async def stype_cb(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.username): return
    parts=cb.data.split("_"); uid=cb.from_user.id; st=get_st(uid)
    ds=st.get("session_ds","?"); name=st.get("session_name","?"); stype="_".join(parts[1:])
    tm_map={"free_first":("🆓 Бесплатная первичная","free"),
            "paid_first":("💼 Платная первичная","paid"),
            "paid_repeat":("💼 Платная повторная","paid")}
    label,ctype=tm_map.get(stype,(stype,"paid"))
    appts=db_get("appts",{})
    if ds not in appts: appts[ds]=[]
    appts[ds].append({"user_id":0,"name":name,"time":"00:00","username":"—","type":ctype,
                       "platform":"—","duration":"—","desc":"Ручная сессия","source":"вручную",
                       "status":"проведена","confirmed":True,"rem_client":True,"rem_admin":True,
                       "session_asked":True,"cli_confirmed":True,"paid":False,"bank":""})
    db_set("appts",appts); clr_st(uid)
    today=date.today()
    await eoa(cb,f"✅ Сессия добавлена.\n{fmt_date(ds)} | {label} | {name}",
              kb=cal_admin(today.year,today.month))
 
@dp.callback_query(F.data.startswith("ban_"))
async def ban_cb(cb: types.CallbackQuery):
    uid=int(cb.data[4:]); bl=db_get("blocked",[])
    if uid not in bl: bl.append(uid); db_set("blocked",bl)
    await eoa(cb,cb.message.text+"\n\n🚫 *Заблокирован*")
 
@dp.callback_query(F.data.startswith("skip_"))
async def skip_cb(cb: types.CallbackQuery):
    await eoa(cb,cb.message.text+"\n\n✅ *Оставлено*")
 
# ═══════════════════════════════════════════════════
# HANDLE_TEXT И MAIN
# ═══════════════════════════════════════════════════
 
@dp.message()
async def handle_text(msg: types.Message):
    uid=msg.from_user.id; text=msg.text.strip() if msg.text else ""
    if is_blocked(uid): return
    if is_flood(uid): return
    reg_user(uid,msg.from_user.username,msg.from_user.first_name)
    if text.startswith("/"):
        clr_st(uid); is_adm=is_admin(msg.from_user.username)
        await msg.answer("🏠 Главное меню:",reply_markup=kb_main(is_adm=is_adm)); return
    if is_admin(msg.from_user.username):
        chats=db_get("admin_chats",[])
        if uid not in chats: chats.append(uid); db_set("admin_chats",chats)
    if has_bad(text):
        cnt=viol(uid,"bad"); await msg.answer("Пожалуйста, давайте общаться без грубостей 🙏")
        if cnt>=5: await _auto_block(uid,f"мат ({cnt}x)"); return
        bk=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🚫 ЗАБЛОКИРОВАТЬ",callback_data=f"ban_{uid}")],
            [InlineKeyboardButton(text="✅ ОСТАВИТЬ",     callback_data=f"skip_{uid}")],
        ])
        await notify_adm(f"⚠️ *Мат* ({cnt}/5)\n@{msg.from_user.username or '?'}(ID:{uid})\n{text[:200]}",kb=bk)
        return
    st=get_st(uid); step=st.get("step","")
 
    if step in ("desc","desc_retry"):
        if len(text)<35:
            st["desc"]=text; set_st(uid,st)
            bk=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ ДОПОЛНИТЬ",         callback_data="desc_extend")],
                [InlineKeyboardButton(text="✅ ОСТАВИТЬ КАК ЕСТЬ", callback_data="desc_keep")],
            ])
            await msg.answer(
                "Вижу, написал совсем немного. Если не хочешь писать - это абсолютно нормально. "
                "Если случайно отправил - можешь дополнить.",reply_markup=bk); return
        name=msg.from_user.first_name or "Клиент"
        st["name"]=name; st["desc"]=text; st["step"]="platform"; set_st(uid,st)
        await msg.answer(f"Через что созвонимся? Выбери удобный вариант 👇{W}",
                         reply_markup=kb_platform("free",depth=3)); return
 
    if step=="adm_fb_text":
        key=st.get("fb_key",""); pf=db_get("pending_feedback",{}); item=pf.get(key,{})
        cli_uid=item.get("uid",0)
        if cli_uid:
            try:
                cname=item.get("name","")
                await bot.send_message(cli_uid,
                    f"Привет, {cname}!\n\nКак ты после нашей встречи? Надеюсь, стало немного яснее.\n\n"
                    f"Как обещал - обратная связь и план работы.\n\n{text}")
                await asyncio.sleep(0.5)
                await bot.send_message(cli_uid,"Если захочешь продолжить работу - я здесь. Вот условия:",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(text="📅🖊️ ЗАПИСАТЬСЯ НА ПЛАТНУЮ",callback_data="paid_info")
                    ]]))
                ft=db_get("feedback_timer",{})
                ft[key]={"uid":cli_uid,"sent":False,"send_at":(datetime.now()+timedelta(hours=72)).isoformat()}
                db_set("feedback_timer",ft)
                item["sent"]=True; pf[key]=item; db_set("pending_feedback",pf)
                today=date.today()
                await msg.answer("✅ Отправлено клиенту.",reply_markup=cal_admin(today.year,today.month))
            except Exception as e: await msg.answer(f"❌ Ошибка: {e}")
        clr_st(uid); return
 
    if step=="adm_manual_tg":
        tg=text.lstrip("@"); st["manual_tg"]=tg
        if st.get("book_ds") and st.get("book_tm"):
            st["step"]="adm_manual_plat"; set_st(uid,st)
            await msg.answer("Платформа для связи?",reply_markup=kb_platform_adm("adm_book_menu",depth=2))
        else:
            st["step"]="adm_book_date"; set_st(uid,st); today=date.today()
            await msg.answer(f"Записываю @{tg}. Выбери дату:",reply_markup=cal_admin(today.year,today.month))
        return
 
    if step=="adm_move_new_time":
        if not re.match(r"^([01]\d|2[0-3]):[0-5]\d$",text):
            await msg.answer("❌ Формат: `14:00`",parse_mode="Markdown"); return
        appts=db_get("appts",{}); old_ds=st["move_ds"]; new_ds=st["move_new_ds"]
        old_tm=st["move_tm"]; target=st["move_uid"]; moved=False
        for rec in appts.get(old_ds,[]):
            if rec["user_id"]==target and rec["time"]==old_tm:
                old_dur=rec.get("duration","30 мин")
                appts[old_ds].remove(rec); _block_adj(old_ds,old_tm,old_dur,unblock=True)
                rec["time"]=text; rec["rem_client"]=False; rec["rem_admin"]=False
                if new_ds not in appts: appts[new_ds]=[]
                appts[new_ds].append(rec); _block_adj(new_ds,text,old_dur); moved=True
                try: await bot.send_message(target,
                        f"📅 Твоя встреча перенесена:\n*{fmt_date(new_ds)}* в *{text}* МСК",
                        parse_mode="Markdown")
                except: pass
                break
        if moved: db_set("appts",appts)
        clr_st(uid); today=date.today()
        await msg.answer("✅ Перенесено." if moved else "❌ Не удалось.",
                         reply_markup=cal_admin(today.year,today.month)); return
 
    if step=="adm_session_name":
        st["session_name"]=text; st["step"]="adm_session_type"; set_st(uid,st)
        kb=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🆓 БЕСПЛАТНАЯ ПЕРВИЧНАЯ", callback_data="stype_free_first")],
            [InlineKeyboardButton(text="💼 ПЛАТНАЯ ПЕРВИЧНАЯ",    callback_data="stype_paid_first")],
            [InlineKeyboardButton(text="💼 ПЛАТНАЯ ПОВТОРНАЯ",    callback_data="stype_paid_repeat")],
        ])
        await msg.answer("Тип встречи?",reply_markup=kb); return
 
    if step=="adm_add_slots_to":
        times_raw=[t.strip() for t in text.split(",")]
        valid=[t for t in times_raw if re.match(r"^([01]\d|2[0-3]):[0-5]\d$",t)]
        if not valid: await msg.answer("❌ Формат: `10:00, 11:00`",parse_mode="Markdown"); return
        ds=st["adm_date"]; slots=db_get("slots",{})
        if ds not in slots: slots[ds]=[]
        for t in valid:
            if t not in slots[ds]: slots[ds].append(t)
        slots[ds]=sorted(slots[ds]); db_set("slots",slots); clr_st(uid)
        await msg.answer(f"✅ Добавлено на {fmt_date(ds)}: {', '.join(valid)}",
                         reply_markup=slots_day_admin(ds)); return
 
    if step=="adm_search_client":
        query=text.lower(); forced=st.get("forced_type","paid")
        ud=db_get("all_users_data",{}); appts=db_get("appts",{}); found={}
        for info in ud.values():
            u=info.get("username",""); n=info.get("name","").lower()
            if (query in n or query in u.lower()) and u not in found:
                found[u]={"name":info.get("name","—"),"username":u,"uid":info.get("uid",0)}
        for ds2,recs in appts.items():
            for r in recs:
                u=r.get("username","")
                if (query in r.get("name","").lower() or query in u.lower()) and u not in found:
                    found[u]={"name":r.get("name","—"),"username":u,"uid":r.get("user_id",0)}
        if not found:
            await msg.answer("Не найдено.",reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ ВВЕСТИ ВРУЧНУЮ",callback_data=f"manual_client_{forced}")],
            ]+nav_admin("adm_back",depth=1))); return
        rows=[[InlineKeyboardButton(
            text=f"👤 {u['name']} @{u['username']}" if u['name'] else f"👤 @{u['username']}",
            callback_data=f"pick_client_{forced}_{u['username']}")]
            for u in found.values()]
        rows+=nav_admin("adm_back",depth=1)
        await msg.answer(f"Найдено {len(found)}:",reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)); return
 
    if step=="adm_reg_username":
        st["reg_username"]=text.lstrip("@"); st["step"]="adm_reg_name"; set_st(uid,st)
        await msg.answer("Введи имя клиента:"); return
    if step=="adm_reg_name":
        st["reg_name"]=text; st["step"]="adm_reg_day"; set_st(uid,st)
        days=["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]
        kb=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=d,callback_data=f"regday_{d}")] for d in days])
        await msg.answer("День недели:",reply_markup=kb); return
    if step=="adm_reg_time":
        if not re.match(r"^([01]\d|2[0-3]):[0-5]\d$",text):
            await msg.answer("❌ Формат: `15:00`",parse_mode="Markdown"); return
        regs=db_get("regulars",[]); regs.append({
            "username":st["reg_username"],"name":st["reg_name"],"day":st["reg_day"],"time":text})
        db_set("regulars",regs); clr_st(uid)
        today=date.today()
        await msg.answer(f"✅ Добавлен @{st['reg_username']}.",reply_markup=cal_admin(today.year,today.month)); return
    if step=="adm_reg_del_name":
        u=text.lstrip("@"); regs=db_get("regulars",[])
        db_set("regulars",[r for r in regs if r["username"].lower()!=u.lower()])
        clr_st(uid); today=date.today()
        await msg.answer(f"✅ @{u} удалён.",reply_markup=cal_admin(today.year,today.month)); return
    if step=="adm_unblock":
        try:
            target=int(text); bl=db_get("blocked",[])
            if target in bl: bl.remove(target); db_set("blocked",bl)
            await msg.answer(f"✅ {target} разблокирован.")
        except: await msg.answer("❌ Введи числовой ID.")
        clr_st(uid); return
    if step=="adm_broadcast_text":
        users=db_get("users",[]); sent=0
        for u in users:
            try: await bot.send_message(u,text); sent+=1; await asyncio.sleep(0.05)
            except: pass
        clr_st(uid); today=date.today()
        await msg.answer(f"✅ Разослано {sent} пользователям.",reply_markup=cal_admin(today.year,today.month)); return
    if step=="adm_post_link":
        users=db_get("users",[]); sent=0
        for u in users:
            try:
                await bot.send_message(u,
                    f"На канале много полезного - вот свежий материал:\n{text}\n\nПодписывайся 👉 @kasikov_psy")
                sent+=1; await asyncio.sleep(0.05)
            except: pass
        clr_st(uid); today=date.today()
        await msg.answer(f"✅ Пост разослан {sent} пользователям.",reply_markup=cal_admin(today.year,today.month)); return
 
    log_act(uid,msg.from_user.username,f"msg:{text[:30]}")
    is_adm=is_admin(msg.from_user.username)
    await msg.answer("Выбери действие:",reply_markup=kb_main(is_adm=is_adm))
 
 
async def main():
    log.info("="*50)
    log.info("БОТ КАСИКОВА v11.0")
    log.info("/admin - панель | /stop - остановить")
    log.info("="*50)
    asyncio.create_task(bg_loop())
    await dp.start_polling(bot)
 
 
if __name__ == "__main__":
    asyncio.run(main())
